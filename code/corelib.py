import os
import pandas as pd
import numpy as np
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
import math
exp = math.exp
from numba import jit
from scipy.optimize import minimize, basinhopping, differential_evolution
from scipy import stats
from os.path import isfile, join
import simplejson as json
import datetime
import copy
import py7zr
import warnings
from dataread import read_spec_json_raw
from datawritter import write_dataframe_as_csv
from utility import communicator

warnings.filterwarnings('ignore')
np.seterr(divide='ignore', invalid='ignore')
path_to_core = "D:\\Core_Archive"
pval_config = {
    1:0.0625,
    2:0.125,
    3:0.25,
    4:0.5,
    5:1,
    6:1.5,
    7:2,
    8:2.5,
    9:3,
    10:3.5,
}
SCALE_FACTOR_IGNORE_T2 = 0.000000000000000001

def combine_array(a1, a2):
    if(a1.shape[0] == 0):
        return a2
    elif (a2.shape[0] == 0):
        return a1
    else:
        return np.c_[a1,a2]

def concat_trans(X, Y):
    output = []
    for i in X:
        for j in Y:
            output.append(i + "#" + j)
    return output

def read_spec_json(json_path, status_path):
    spec = read_spec_json_raw(json_path)    
    if 'stochastic_sea' not in spec["basic_settings"].keys():
        spec["basic_settings"]['stochastic_sea'] = 0
    else:
        spec["basic_settings"]['stochastic_sea'] = float(spec["basic_settings"]['stochastic_sea'])

    if 'agg_period_search' not in spec["basic_settings"].keys():
        spec["basic_settings"]['agg_period_search'] = 1
    else:
        spec["basic_settings"]['agg_period_search'] = float(spec["basic_settings"]['agg_period_search'])

    if 'agg_constraint_cutoff' not in spec["basic_settings"].keys():
        spec["basic_settings"]['agg_constraint_cutoff'] = 1
    else:
        spec["basic_settings"]['agg_constraint_cutoff'] = float(spec["basic_settings"]['agg_constraint_cutoff'])
    
    #Custom dependent 
    if 'is_Custom_Dept_Agg' not in spec["basic_settings"].keys():
        spec["basic_settings"]['is_Custom_Dept_Agg'] = 0
    else:
        spec["basic_settings"]['is_Custom_Dept_Agg'] = int(spec["basic_settings"]['is_Custom_Dept_Agg'])
    
    if 'custom_Dept_Agg_Val' not in spec["basic_settings"].keys():
        spec["basic_settings"]['custom_Dept_Agg_Val'] = 0.0
    else:
        spec["basic_settings"]['custom_Dept_Agg_Val'] = float(spec["basic_settings"]['custom_Dept_Agg_Val'])
    
    if 'isRandomiziedFilter' not in spec["basic_settings"].keys():
        spec["basic_settings"]['isRandomiziedFilter'] = 0
    else:
        spec["basic_settings"]['isRandomiziedFilter'] = int(spec["basic_settings"]['isRandomiziedFilter'])
    
    #===============================================================
    if 'single_transformation' not in spec["basic_settings"].keys():
        spec["basic_settings"]['single_transformation'] = False
    else:
        spec["basic_settings"]['single_transformation'] = bool(spec["basic_settings"]['single_transformation'])

    if 'exclude_dependent_zero' not in spec["basic_settings"].keys():
        spec["basic_settings"]['exclude_dependent_zero'] = False
    else:
        spec["basic_settings"]['exclude_dependent_zero'] = bool(spec["basic_settings"]['exclude_dependent_zero'])

    if 'fast_opt' not in spec["basic_settings"].keys():
        spec["basic_settings"]['fast_opt'] = False
    else:
        spec["basic_settings"]['fast_opt'] = bool(spec["basic_settings"]['fast_opt'])

    if 'old_model_path' not in spec["basic_settings"].keys(): 
        spec["basic_settings"]["old_model_path"] = False
    elif spec["basic_settings"]["old_model_path"]:
        spec["basic_settings"]["old_model_path"] = "\\".join([path_to_core, spec["basic_settings"]["old_model_path"]])
    else:
        spec["basic_settings"]["old_model_path"] = False

    if 'seasonality_exclude' not in spec["basic_settings"].keys(): 
        spec["basic_settings"]["seasonality_exclude"] = []
    
    if spec["basic_settings"]['isRandomiziedFilter']==1:
        randomizedList = spec["basic_settings"]["randomize_level"]
        randomizedList = pd.DataFrame(randomizedList).to_numpy().T
        randomized_layers = dict(zip(randomizedList[1], randomizedList[0]))
    else: 
        randomized_layers = {} 
    if 'is_tolerance_level' not in spec["basic_settings"].keys(): 
        spec["basic_settings"]["is_ignore_T2"] = False
    else:
        spec["basic_settings"]["is_ignore_T2"] = bool(spec["basic_settings"]["is_tolerance_level"])   

    basic_settings = spec["basic_settings"]
    basic_settings["include_seasonality"] = bool(basic_settings["include_seasonality"])
    basic_settings["include_intercept"] = bool(basic_settings["include_intercept"])
    basic_settings["pval_config"] = {int(k): v for k,v in basic_settings["pval_config"].items()}

    configuration = spec["configuration"]
    dependent_name = basic_settings["dependent_colname"]

    reporting_period = basic_settings["reporting_period"]
    reporting_period = pd.DataFrame(reporting_period)
    reporting_period.set_index('name', inplace=True)

    configuration = pd.DataFrame(configuration)
    configuration["include"] = configuration["include"].astype(bool)
    configuration = configuration[configuration["include"]]

    contribution = configuration[["variable", "contri_period", "min", "max", "sign"]]
    # contribution["min"] = contribution["min"] / 100
    # contribution["max"] = contribution["max"] / 100
    is_volumetric_constraints = False
    if "isVolumetricConstraint" in configuration:
        contribution["isVolumetricConstraint"] = configuration["isVolumetricConstraint"]
        is_volumetric_constraints = contribution['isVolumetricConstraint'].eq(1).any()
    print("is_volumetric_constraints: ", is_volumetric_constraints)
    contribution.set_index('variable', inplace=True)
    contribution = contribution.T.to_dict()

    # The condition is so that old specs with no postmul in json works
    if "postmul" in configuration.columns:
        postmul = configuration[["variable", "postmul"]]
        postmul.set_index('variable', inplace=True)
        postmul = postmul[~postmul["postmul"].isnull()].squeeze(axis=1).to_dict()
        for k, v in postmul.items():
            if str(v).isnumeric():
                postmul[k] = float(v)
            elif "|" in v:
                postmul[k] = v.split("|")
    else:
        postmul = {}
    nontransformation_cols = list(configuration[configuration["transformation_values"].isnull()]["variable"].sort_values())
    transformations = configuration[~configuration["transformation_values"].isnull()]
    transformations = transformations[["transformation","variable", "transformation_values"]].sort_values(["transformation","variable"])
    transformations.set_index('variable', inplace=True)
    transformations = transformations.T.to_dict()
    
    var_wise_tolerance_level_dict = {}
    # if spec["basic_settings"]["is_ignore_T2"]: 
    if "tolerance_level" in configuration:
        var_wise_tolerance_level = configuration[configuration["tolerance_level"]!=False]
        var_wise_tolerance_level = var_wise_tolerance_level[["tolerance_level","variable", "toleranceValue"]].sort_values(["toleranceValue","variable"])
        var_wise_tolerance_level.set_index('variable', inplace=True)
        var_wise_tolerance_level = var_wise_tolerance_level.T.to_dict()
        
        for i in var_wise_tolerance_level:
            var_wise_tolerance_level_dict[i] = {contribution[i]["contri_period"]:var_wise_tolerance_level[i]["toleranceValue"]}
            # var_wise_tolerance_level_dict[i][contribution[i]["contri_period"]] = var_wise_tolerance_level[i]["toleranceValue"]
        print("var_wise_tolerance_level: ", var_wise_tolerance_level)
    transformation_cols = list(transformations.keys())
    transformation_strings = []
    # exit()
    # Generate Trnasformation Combinations -----------> # Difficulty Level 4
    initial_pval = {}
    try: 
        for variable, value in transformations.items():
            trans_varnames = []
            if(value["transformation"] == None):
                raise Exception("102| " + "No Transformation for " + variable)
            transformations = value["transformation"].split("#")
            transformation_values = str(value["transformation_values"]).split("#")
            trans_string = {}
            for i in range(len(transformations)):
                tranformation = transformations[i]
                values = transformation_values[i]
                if(tranformation == "PVAL"):
                    trans_string["PVAL"] = []
                    values = values.split("|")
                    if values[0][0] == "{":
                        split = values[0].split("}")
                        initial_pval[variable] = split[0][1:]
                        values[0] = split[1]
                    for l in values[0].split(","):
                        for d in values[1].split(","):
                            trans_string["PVAL"].append("PVAL" + "_" + l + "|" + d)
                elif(tranformation == "APL"):
                    trans_string["APL"] = []
                    values = values.split("|")
                    for a in values[0].split(","):
                        for p in values[1].split(","):
                            for l in values[2].split(","):
                                trans_string["APL"].append("APL" + "_" + a + "|" + p + "|" + l)
                elif(tranformation == "RA"):
                    trans_string["RA"] = []
                    for w in values.split(","):
                        trans_string["RA"].append("RA" + "_" + w)
                elif(tranformation == "LAG"):
                    trans_string["LAG"] = []
                    for w in values.split(","):
                        trans_string["LAG"].append("LAG" + "_" + w)
                else:
                    raise Exception("101|" + tranformation + " Unrecognised for variable " + variable)
            trans_used = list(trans_string.keys())
            output = trans_string[trans_used[0]]
            for i in range(1, len(trans_used)):
                output = concat_trans(output, trans_string[trans_used[i]])
            transformation_strings += [variable + "__" + trans for trans in output]
    except Exception as e: 
        raise Exception("103| Transformation error for variable " + variable)

    if(len(transformation_strings) - len(set(transformation_strings)) > 0):
        diff_vars = list(set([x for x in transformation_strings if transformation_strings.count(x) > 1]))
        diff_vars = list(set([x.split("__")[0] for x in diff_vars]))
        raise Exception("107| Transformation workflow values are generating duplicates for variables " + str(diff_vars))
    # No of variables
    var_count = len(transformation_cols) + len(nontransformation_cols)
    if basic_settings["include_seasonality"]:
        var_count += 1
    if basic_settings["include_intercept"]:
        var_count += 1
    if basic_settings["stochastic_sea"]:
        var_count += 1

    # For Fast Coorelation transformation_strings
    transformation_sep = [s.split("__")[0] for s in transformation_strings]
    l = len(transformation_strings)
    same_check = np.zeros((l,l))
    for i in range(l):
        for j in range(i, l):
            if transformation_sep[j] != transformation_sep[i]:
                same_check[j][i] = 1

    same_check = same_check.astype(np.bool)
    # Handle post mul cols
    all_cols_read = transformation_cols + nontransformation_cols + [dependent_name]
    postmul_values = list(list(j) if not isinstance(j, str) else j for j in set([tuple(i) if not isinstance(i, str) else i for i in postmul.values()]))#list(set(postmul.values()))
    postmul_only_col = []
    for col in postmul_values:
        if isinstance(col, str) and col != "__DEPSEA" and col not in all_cols_read:
            postmul_only_col.append(col)
        elif isinstance(col, list):
            for v in col:
                if v not in all_cols_read:
                    if v not in postmul_only_col:
                        postmul_only_col.append(v)
    all_cols_read += postmul_only_col
    # Handle Splits and Weekly Seasonality - Remove them from reading from ADS
    split_variables = []
    all_cols_read_final = []
    only_split = []
    dummy = []
    for col in all_cols_read:
        if "#WEEKLY_" in col or "#DUMMY(" in col:
            dummy.append(col)
        elif "(" in col:
            split_variables.append(col)
            _split = col.split("(")
            var = _split[0]
            if var not in all_cols_read and var not in only_split:
                only_split.append(var)
            period = _split[1][:-1]
        else:
            all_cols_read_final.append(col)
    all_cols_read = all_cols_read_final + only_split

    # Handle Seasonality Variable Based
    if(basic_settings["include_seasonality"] == 0 and basic_settings["seasonality_variable"] != "" and basic_settings["seasonality_variable"] != "#WEEKLY"):
        if basic_settings["seasonality_variable"] not in all_cols_read:
            all_cols_read = all_cols_read + [basic_settings["seasonality_variable"]]
    nontransformation_cols = custom_dummy_sort(nontransformation_cols)

    spec = {
        "same_check": same_check,
        "dependent_name": dependent_name,
        'contribution_range': contribution,
        'var_count': var_count,
        'config': basic_settings,
        'reporting_period': reporting_period,
        "transformation_sep": transformation_sep,
        "transformed_cols": transformation_strings, # Big_Array
        "transformation_cols": transformation_cols,
        "nontransformation_cols": nontransformation_cols,
        "all_cols_read": all_cols_read,
        'postmul': postmul,
        "initial_pval": initial_pval,
        "split_variables": split_variables,
        "dummy": dummy,
        "old_model_path": spec["basic_settings"]["old_model_path"],
        "exclude_dependent_zero": spec["basic_settings"],
        "randomized_layers": randomized_layers,
        "var_wise_tolerance_level_dict": var_wise_tolerance_level_dict,
        "is_ignore_T2": spec["basic_settings"]["is_ignore_T2"],
        "is_volumetric_constraints": is_volumetric_constraints
    }
    return spec

def custom_dummy_sort(var_list):
    dummy = []
    other = []
    for col in var_list:
        if("#" in col):
            dummy.append(col)
        else:
            other.append(col)
    dummy.sort(key=len) # This is 2nd level sort - 1st it is already sorted by lexicographically
    return other + dummy # Put dummy to the end

def spec_xlsx_to_json(xlsx_path):
    # Read Spec
    base_hypothesis = pd.read_excel(xlsx_path, sheet_name='Base Hypothesis', header=None, parse_dates=True,date_parser=lambda x: pd.to_datetime(x).strftime("%m/%d/%Y"))
    model_specifications = pd.read_excel(xlsx_path, sheet_name='Model Specifications', header=None, parse_dates=True,date_parser=lambda x: pd.to_datetime(x).strftime("%m/%d/%Y"))
    config_sheet = pd.read_excel(xlsx_path, sheet_name='Config', header=None, parse_dates=True,date_parser=lambda x: pd.to_datetime(x).strftime("%m/%d/%Y"))
    # Read Config
    seasonality = config_sheet[1][39].strip().upper()
    seasonality_variable = ""
    # include_seasonality is for dependend based, while seasonality_variable is for variable base, with exception #WEEKLY for weekly dummy
    if seasonality == "WEEKLY":
        include_seasonality = 0
        seasonality_variable = "#WEEKLY"
    elif seasonality == "YES" or seasonality == "SG0M":
        include_seasonality = 1
    elif seasonality == "NO" or seasonality == "":
        include_seasonality = 0
    else : # Variable Based
        include_seasonality = 0
        seasonality_variable = config_sheet[1][39].strip()

    basic_settings = {
        "period_colname": config_sheet[0][10].strip(),
        "geo_colname": config_sheet[0][12].strip(),
        "dependent_colname": base_hypothesis[1][15].strip(),

        "include_seasonality": include_seasonality,
        "seasonality_variable": seasonality_variable, # Empty string means dependent based, else variable based or weekly.
        "seasonality_cycle": 52,
        "weekly_cycle": 52,
        "include_intercept": int(config_sheet[2][55].strip().upper() == "YES"),

        "transformation_period": base_hypothesis[0][3],
        "modelling_period": {
            "start_date": base_hypothesis[1][3],
            "end_date": base_hypothesis[2][3],
        },
        "pval_config": pval_config
    }

    # Read Base Hypothesis
    reporting_period = base_hypothesis.iloc[2:50,4:7]
    end_trim = reporting_period[6].to_list().index(np.nan)
    reporting_period = reporting_period[:end_trim]
    reporting_period.columns = reporting_period.iloc[0].str.upper().str.strip()
    reporting_period.dropna(how="all", inplace=True)
    reporting_period = reporting_period.drop(reporting_period.index[0])
    reporting_period = reporting_period.reset_index(drop = True)
    reporting_period = reporting_period.rename(columns={"SHORT NAME": "name", "START DATE": "start_date", "END DATE": "end_date"}, errors="raise")
    reporting_period["name"] = reporting_period["name"].astype(str)

    shortnames = reporting_period["name"].to_list()    

    # reporting_period = list(reporting_period.T.to_dict().values())
    reporting_period = reporting_period.T   
    reporting_period.loc["name"] = shortnames
    reporting_period = list(reporting_period.to_dict().values())
    basic_settings["reporting_period"] = reporting_period

    contribution_range = base_hypothesis.iloc[2:, 10:].T
    contribution_range[3] = contribution_range[3].str.split(' \(', expand=True)[0].str.upper()
    contribution_range[2] = contribution_range[2].ffill().astype(str)
    contribution_range = contribution_range.set_index([2,3], drop=True)
    contribution_range.columns = contribution_range.iloc[0]
    contribution_range = contribution_range.drop(contribution_range.index[0])
    contribution_range = contribution_range.replace(r'^\s*$', np.nan, regex=True).fillna(0).T
    contribution_range = contribution_range[~contribution_range.index.isnull()]
    range_dict = {}
    # For Collective Error Message
    stop = False
    min_max_err = ""
    mul_err = ""
    for name in shortnames: # CY YA Y2A etc
        contri_for_name = contribution_range[name]
        contri_for_name = contri_for_name[(contri_for_name["MAX"] != 0) | (contri_for_name["MIN"] != 0)]
        for variable, row in contri_for_name.iterrows():
            if row["MIN"] > row["MAX"]:
                stop = True
                min_max_err += str(variable) + ", "
            if variable in range_dict.keys():
                stop = True
                mul_err += str(variable) + ", "
                continue
            range_dict[variable] = {
                "contri_period": name,
                "min": row["MIN"],
                "max": row["MAX"],
            }
    
    if stop:
        msg = ""
        if min_max_err != "":
            msg += min_max_err[:-2] + " has min > max ||"
        if mul_err != "":
            msg += mul_err[:-2] + " has multiple contribution"

        communicator("ERROR| " + msg, xlsx_path.split(".")[0] + "_error.txt")
        raise Exception(msg)

    no_contribution_variables = list(contribution_range[(contribution_range.T == 0).all()].index)
    for variable in no_contribution_variables:
        if variable in range_dict.keys():
                communicator("ERROR| " + variable + " is repeated", xlsx_path.split(".")[0] + "_error.txt")
                raise Exception(variable + " is repeated")
        range_dict[variable] = {
            "contri_period": np.nan,
            "min": 0,
            "max": 0,
        }

    pd_contri = pd.DataFrame.from_dict(range_dict).T
    # Read Model Specifications
    new_header = model_specifications.iloc[3].str.strip().str.upper() #grab the first row for the header
    model_specifications = model_specifications[5:] #take the data less the header row
    model_specifications.columns = new_header #set the header row as the df header
    model_specifications.set_index('VARIABLES', inplace=True)
    model_specifications["WORKFLOW"] = model_specifications["WORKFLOW"].str.replace("PLN\|PDY", "PVAL")
    model_specifications["WORKFLOW"] = model_specifications["WORKFLOW"].str.replace("ADS\|APWR\|ALN", "APL")

    model_specifications["postmul"] = model_specifications["POST MULTIPLICATION"].replace(r'^\s*$', np.NaN, regex=True)

    model_specifications["SIGN"] = model_specifications["SIGN"].str.upper()
    try:
        configuration = pd.concat([model_specifications, pd_contri],sort=True, axis=1)
    except Exception as e:
        if not(model_specifications.index.is_unique and pd_contri.index.is_unique):
            communicator("ERROR| " + "Check for duplications in model specification or contribution variables", xlsx_path.split(".")[0] + "_error.txt")
            raise Exception("Check for duplications in model specification or contribution variables")
        diff =  set(pd_contri.T.keys().array).difference(set(model_specifications.T.keys().array))
        communicator("ERROR| " + "Varibles present in contibution must be present in specification. Check the Following - " + repr(diff), xlsx_path.split(".")[0] + "_error.txt")
        raise Exception("Varibles present in contibution must be present in specification. Check the Following - " + repr(diff))
    # Check for split variables
    for col in pd_contri.T.keys().array:
        if "(" in col:
            period = col.split("(")[1][:-1]
            if (not period in shortnames):
                raise Exception(col + " " + period + " did not match any period.")
    
    configuration["INCLUDE VARIABLE"] = (configuration["INCLUDE VARIABLE"].str.upper() == "YES").astype(int)
    configuration = configuration.rename(columns={"TYPE": "type", "SIGN": "sign", "INCLUDE VARIABLE":"include", "WORKFLOW": "transformation", "WORKFLOW VALUES": "transformation_values", "BUCKET": "desp", }, errors="raise")
    variable_list = configuration.index
    configuration['variable'] = variable_list
    configuration = configuration[['include', 'variable', 'type', 'desp', 'transformation', 'transformation_values', 'postmul', "sign", "contri_period", "min", "max" ]]
    # configuration["min"] = configuration["min"] * 100
    # configuration["max"] = configuration["max"] * 100
    try:
        transformation_with_vals = configuration[configuration['transformation_values'].notnull()]
        transformation_with_no_strings_for_vals = transformation_with_vals[transformation_with_vals['transformation'].isnull()]
        transformation_with_strings = configuration[configuration['transformation'].notnull()]
        transformation_with_no_vals_for_strings = transformation_with_strings[transformation_with_strings['transformation_values'].isnull()]   
        if len(list(transformation_with_no_strings_for_vals.T.to_dict().values())) > 0 :
            variable = ",".join([i['variable'] for i in transformation_with_no_strings_for_vals.T.to_dict().values()])
            raise Exception("Spec|Transformation error for variable " + variable)
        if len(list(transformation_with_no_vals_for_strings.T.to_dict().values())) > 0 :
            variable = ",".join([i['variable'] for i in transformation_with_no_vals_for_strings.T.to_dict().values()])
            raise Exception("Spec|Transformation error for variable " + variable)    
    except Exception as e:
        communicator("ERROR| " + "Spec| Transformation error for variable " + variable, xlsx_path.split(".")[0] + "_error.txt")
        raise Exception("Spec| Transformation error for variable " + variable)
    configuration = list(configuration.T.to_dict().values())
    
    if seasonality_variable == "#WEEKLY":
        for i in range(1, basic_settings["weekly_cycle"] + 1):
            variable = "#WEEKLY_" + str(i) + "|1|" + str(basic_settings["weekly_cycle"])
            if variable in variable_list:
                continue
            desp = "Weekly Dummy " + str(i)
            configuration.append({'include': 1, 'variable': variable, 'type': 'Incremental', 'desp': desp, 'transformation': np.nan, 'transformation_values': np.nan, 'postmul': np.nan, 'sign': 'ANY', 'contri_period': np.nan, 'min': 0.0, 'max': 0.0})
    # Write to Json
    spec = {
        "basic_settings": basic_settings,
        "configuration": configuration,
    }
    json_path = xlsx_path.split(".x")[0] + ".json"
    with open(json_path, 'w') as fp:
        spec_json = json.dumps(spec, ignore_nan=True, default=datetime.datetime.isoformat)
        fp.write(spec_json)
    fp.close()
    # json.dumps(spec, fp, default=datetime_handler, ignore_nan=True)
    return json_path

def split_variables(geo_df, spec):
    split_variables = spec["split_variables"]
    reporting_period_index = spec["reporting_period_index"]
    for col in split_variables:
        _split = col.split("(")
        variable = _split[0]
        period = _split[1][:-1]
        loc = reporting_period_index[period]
        variable_data = geo_df[variable]
        new_variable_data = np.zeros_like(variable_data)
        new_variable_data[loc[0]:loc[1] + 1] = variable_data[loc[0]:loc[1] + 1]
        geo_df[col] = new_variable_data
    return geo_df

def gen_dummy(spec):
    dummy = spec["dummy"]
    data_l = spec['week_count']
    t_data_l = spec["data_week_count"]
    reporting_period_index = spec["reporting_period_index"]
    dummy_data = {}
    for col in dummy:
        if("#WEEKLY" in col):
            params = col.split("_")[1].split("|")
            [start, howmany, cycle] = params
            new_variable_data = np.zeros(data_l)
            for i in range(int(start) - 1, data_l, int(cycle)):
                new_variable_data[i: i + int(howmany)] = 1
            dummy_data[col] = new_variable_data
        elif("#DUMMY(" in col):
            _split = col.split("(")
            variable = _split[0]
            period = _split[1][:-1]
            loc = reporting_period_index[period]
            new_variable_data = np.zeros(data_l)
            new_variable_data[loc[0]:loc[1] + 1] = 1
            dummy_data[col] = new_variable_data
        elif("#DUMMY_" in col):
            _split = col.split("_")[1].split("|")
            for params in _split:
                [loc, howmany] = params.split("-")
                loc = int(loc)
                howmany = int(howmany)
                new_variable_data = np.zeros(data_l)
                new_variable_data[loc: loc + howmany] = 1
            dummy_data[col] = new_variable_data
        # Match Length with actual data
        raw_data =  np.zeros(t_data_l)
        raw_data[spec['modelling_start_index']:spec['modelling_end_index'] +1] = dummy_data[col]
        dummy_data[col] = raw_data
    return dummy_data

def bounds_calc(contribution_range, transformed_support_agg, support_agg, dependent_agg, allvar_support_colnames, periods, agg_constraint_cutoff, agg_period_search, contri_override, spec, is_store_need_scaled_bounds = None, dependent_distribution = None):  # -----------> # Difficulty Level 5
    ignore_T2 = spec["is_ignore_T2"]
    var_wise_cutoff = spec["var_wise_tolerance_level_dict"]
    
    contribution_range = copy.deepcopy(contribution_range)
    for item in contri_override:
        variable = item.pop('variable')
        contribution_range[variable] = item
    reverse_period = {}
    system_unbound = []
    for i in range(len(periods)):
        reverse_period[periods[i]] = i

    transformed_support = np.array(transformed_support_agg)
    support_agg = np.array(support_agg)
    sum_support_agg = support_agg.sum(axis=0)
    support_abs = np.abs(support_agg)

    # Swap Period or sign contraint Based of support
    for col_index in range(len(allvar_support_colnames)):
        col = allvar_support_colnames[col_index]
        selected_period = contribution_range[col]["contri_period"]
        try:
            updated_agg_constraint_cutoff = agg_constraint_cutoff if col not in var_wise_cutoff else var_wise_cutoff[col][selected_period]
        except Exception as e:
            raise Exception("Error in variable cutoff: " +col + "|"+str(e))
        if not selected_period or selected_period == "FIXED":
            continue # all zeros

        # pos = reverse_period[selected_period]
        pos = reverse_period[selected_period] if selected_period in reverse_period.keys() else None
        if pos is None:
            raise Exception(f"Reporting period({selected_period}) not found in Spec, please correct your spec.")

        if support_abs[pos][col_index] < agg_period_search:
            support_col_max = support_abs[:,col_index].max()
            if support_col_max < updated_agg_constraint_cutoff:
                if support_col_max != 0: system_unbound.append({"c": col, "r": "T1|" + str(np.round(support_col_max, 2))})
                else: system_unbound.append({"c": col, "r": "S0"})
                l = contribution_range[col]["min"]
                u = contribution_range[col]["max"]
                if (l*u >= 0):
                    if l + u > 0:
                        contribution_range[col]["sign"] = "POS"
                    else :
                        contribution_range[col]["sign"] = "NEG"
                contribution_range[col]["contri_period"] = False
                # if col in var_wise_cutoff:
                #     cut_off_val = var_wise_cutoff[col][selected_period]
                #     var_wise_cutoff[col] = {False: cut_off_val }
            else:
                maxLoc = np.argmax(support_abs[:,col_index])
                contribution_range[col]["contri_period"] = periods[maxLoc]
            # exit()
    # if actual for a period is 0, look for a period where dependent and support is not zero, else it should be sign contraint else with desired volume
    # print("agg_period_search: ", agg_period_search)
    # exit()
    for col_index in range(len(allvar_support_colnames)):

        col = allvar_support_colnames[col_index]
        selected_period = contribution_range[col]["contri_period"]
        try:
            updated_agg_constraint_cutoff = agg_constraint_cutoff if col not in var_wise_cutoff else var_wise_cutoff[col][selected_period]
        except Exception as e:
            raise Exception("Error in variable cutoff: " +col + "|"+str(e))
        # if(col == "ET2_MUL_BRN"):
        #     print("selected_period ", selected_period)
        if not selected_period or selected_period == "FIXED":
            continue # all zeros

        # pos = reverse_period[selected_period]
        pos = reverse_period[selected_period] if selected_period in reverse_period.keys() else None
        if pos is None:
            raise Exception(f"Reporting period({selected_period}) not found in Spec, please correct your spec.")

        if dependent_agg[pos] == 0: # look for a period where dependent and support is not zero .. starting from highest support, if this fails this condition remails -> dependent_agg[pos] == 0
            supports = support_abs[:,col_index]
            support_map = supports.tolist()
            for s in np.sort(supports)[::-1]: # Decending sort
                if s >= agg_constraint_cutoff:
                    check_pos = support_map.index(s)
                    if dependent_agg[check_pos] != 0:
                        pos = check_pos
                        break
                else:
                    break

        dep_agg_period_value = dependent_agg[pos]
        value_transformed = transformed_support[pos][col_index]

        if not ignore_T2 and (dep_agg_period_value == 0 or value_transformed == 0): #ignore_T2  = Neglecting T2 cases for all variables #(not ignore_T2 or value_transformed==0) and
            system_unbound.append({"c": col, "r": "T2|DEP/SUP0"})
            l = contribution_range[col]["min"]
            u = contribution_range[col]["max"]
            if (l*u >= 0):
                if l + u > 0:
                    contribution_range[col]["sign"] = "POS"
                else :
                    contribution_range[col]["sign"] = "NEG"
            contribution_range[col]["contri_period"] = False
        else:
            if ignore_T2:
                    if dep_agg_period_value == 0:
                        dep_agg_period_value = SCALE_FACTOR_IGNORE_T2
                    if value_transformed == 0:
                        value_transformed = SCALE_FACTOR_IGNORE_T2
            if is_store_need_scaled_bounds is None or is_store_need_scaled_bounds:
                if "isVolumetricConstraint" in contribution_range[col] and bool(contribution_range[col]["isVolumetricConstraint"]):
                    if dependent_distribution is None:
                        raise Exception("Please check for volumetric constraints")
                    contribution_range[col]["min"] = (contribution_range[col]["min"]*dependent_distribution)/value_transformed#store count = 14: Avg store
                    contribution_range[col]["max"] = (contribution_range[col]["max"]*dependent_distribution)/value_transformed#store count = 14: Avg store
                else:
                    contribution_range[col]["min"] *= (dep_agg_period_value / value_transformed)
                    contribution_range[col]["max"] *= (dep_agg_period_value / value_transformed)
    # Bound formation -> Sign + support check and min > max check
    bounds = []
    scale_vector = np.ones(len(allvar_support_colnames))
    for col_index in range(len(allvar_support_colnames)):
        col = allvar_support_colnames[col_index]
        selected_period = contribution_range[col]["contri_period"]
        # if(col == "OOH_SPD"):
        #     print(contribution_range[col]["sign"])
        #     print(contribution_range[col]["contri_period"])
        #     print(contribution_range[col]["min"])
        #     print(contribution_range[col]["max"])
        #     print(sum_support_agg)
        # exit()
        if not selected_period:
            sign = contribution_range[col]["sign"]
            sum_support = sum_support_agg[col_index]
            scale_vector[col_index] = 0
            if sign == "POS":
                if sum_support >= 0:
                    bounds.append([0, np.inf])
                else:
                    bounds.append([-np.inf, 0])
            elif sign == "NEG":
                if sum_support >= 0:
                    bounds.append([-np.inf, 0])
                else:
                    bounds.append([0, np.inf])
            else:
                bounds.append([-np.inf, np.inf])
        elif selected_period == "FIXED":
            bounds.append([1, 1])
        else:
            # pos = reverse_period[selected_period]
            pos = reverse_period[selected_period] if selected_period in reverse_period.keys() else None
            if pos is None:
                raise Exception(f"Reporting period({selected_period}) not found in Spec, please correct your spec.")
            # value_transformed = transformed_support[pos][col_index]
            l = contribution_range[col]["min"]
            u = contribution_range[col]["max"]
            # print(u - l)
            if is_store_need_scaled_bounds is None or is_store_need_scaled_bounds:
                scale_vector[col_index] = u - l 
            if(l > u):
                l,u=u,l
            bounds.append([l, u])
        # if(col == "ET2_MUL_BRN"):
    # print(bounds[19])
    # print('bounds: ', bounds)
    # print('scale_vector: ', scale_vector)
    return bounds , scale_vector, system_unbound

def granular_writter(path, file_name, spec, support_optimized, predicted, dependent_col, stats_dict, all_cols, transformed_support, trans_support_no_postmul, corr_data, bounds, scale_factor):
    period_s = spec["period_s"]
    dependent_name = spec['dependent_name']
    # Stats
    if spec['granular_stats']:
        scale_col = np.empty_like(stats_dict['T Stat'])
        all_cols_no_trans = [i.split("__")[0] for i in all_cols]
        for i in range(len(all_cols)):
            if all_cols_no_trans[i] in scale_factor.keys() and scale_factor[all_cols_no_trans[i]] != 1 :
                scale_col[i] = scale_factor[all_cols_no_trans[i]]
            else :
                scale_col[i] = np.nan
        stats_dict['Pval'] = stats.t.sf(np.abs(stats_dict['T Stat']), support_optimized.shape[0] - 1) * 2
        stats_dict['Scale'] = scale_col
        df = pd.DataFrame(stats_dict)
        df.insert(0, "Variable", all_cols)
        stats_out = os.path.join(path, "stats", file_name + '__stats.csv')
        write_dataframe_as_csv(df, stats_out)
        # df.to_csv(path + '/stats/' + file_name + '__stats.csv', index=False)
    # Granular result
    if spec['granular_result']:
        final_file = np.c_[support_optimized, predicted, dependent_col]
        all_cols.append('Predicted')
        all_cols.append(dependent_name)
        df = pd.DataFrame(final_file, columns=all_cols)
        df.insert(0, period_s.name, period_s)
        result_out = os.path.join(path, "granular_result", file_name + '__result.csv')
        write_dataframe_as_csv(df, result_out)
        # df.to_csv(path + '/granular_result/' + file_name + '__result.csv', index=False)
    # Transformed Support
    if spec['granular_result'] and spec['granular_support']:
        t_file = np.c_[transformed_support, dependent_col]
        t_wp_file = np.c_[trans_support_no_postmul, dependent_col] #new added
        all_cols.remove('Predicted') # Remove Predicted
        df = pd.DataFrame(t_file, columns=all_cols)
        df1 = pd.DataFrame(t_wp_file, columns=all_cols)
        df.insert(0, period_s.name, period_s)
        df1.insert(0, period_s.name, period_s)
        # bounds_df = pd.DataFrame(bounds, columns=["MIN", "MAX"], index=all_cols[:-2]).T
        # bounds_df['Predicted'] = [0, 0]
        # bounds_df[dependent_name] = [0, 0]
        # bounds_df[period_s.name] = [0, 0]
        # print(bounds_df)
        # print(df)
        # df = df.append(bounds_df)
        # print(df)
        support_out = os.path.join(path, "granular_support", file_name + '__transformed_support.csv')
        support_out1 = os.path.join(path, "granular_support_no_postmul", file_name + '__transformed_support_no_postmul.csv')
        write_dataframe_as_csv(df, support_out)
        write_dataframe_as_csv(df1, support_out1)
        # df.to_csv(path + '/granular_support/' + file_name + '__transformed_support.csv', index=False)
    if spec['granular_correlation'] and isinstance(corr_data, pd.DataFrame):
        corr_data.to_csv(path + '/granular_result/' + file_name + '__coorelation_stats.csv')

def prep_optimization(file_name, geo_df, spec, dependent_col, selected_transformations, all_transformations):
    # Handle exclude_dependent_zero
    if(file_name in spec['geo_ex_exdep'].keys()):
        exclude_dependent_zero = False
    else:
        exclude_dependent_zero = spec["config"]["exclude_dependent_zero"]

    # Handle Multi Contribution
    geo_map_contri = spec['geo_map_contri']
    contri_map = spec['contri_map']
    global_contri = spec['contribution_range']
    contribution_range = contri_map[geo_map_contri[file_name]] if file_name in geo_map_contri.keys() else global_contri

    reporting_period_index = spec['reporting_period_index']
    # Col names
    transformed_colnames = list(selected_transformations.values())
    transformation_colnames = list(selected_transformations.keys())

    nontransformation_colnames = spec['nontransformation_cols']
    transformed_support_colnames = transformed_colnames + nontransformation_colnames
    allvar_support_colnames = transformation_colnames + nontransformation_colnames

    # 3 Data  - transformed and 2 support, transformed support
    transformed_data = np.array([all_transformations[col] for col in selected_transformations.values()]).T
    support_transformation_data = np.array([geo_df[col] for col in transformation_colnames]).T
    support_nontransformation_data = np.array([geo_df[col] for col in nontransformation_colnames]).T
    transformed_support = combine_array(transformed_data, support_nontransformation_data)
    support = combine_array(support_transformation_data, support_nontransformation_data)
    
    # 3.1 transformed support (copy)
    trans_support_no_postmul = np.copy(transformed_support)
    
    # 4 Calc Seasonality
    seasonality_exclude = spec['config']['seasonality_exclude']
    if(file_name in spec['geo_map_sea'].keys()):
        seasonality_exclude = spec['sea_map'][spec['geo_map_sea'][file_name]]
    if spec['modelling_start_index'] or spec['modelling_end_trim']:
        dependent_col_trim = dependent_col[spec['modelling_start_index']:spec['modelling_end_trim']]
        seasonality_col = gen_seasonality(np.copy(dependent_col_trim), spec['config']['seasonality_cycle'], seasonality_exclude, exclude_dependent_zero)
    else:
        seasonality_col = gen_seasonality(np.copy(dependent_col), spec['config']['seasonality_cycle'], seasonality_exclude, exclude_dependent_zero)

    # 5 Post Multiplication
    postmul = spec['postmul']
    for var in postmul.keys():
        pos = allvar_support_colnames.index(var)
        if isinstance(postmul[var], str):
            if postmul[var] ==  "__DEPSEA" and spec["config"]["include_seasonality"]:
                # Divide Dep / Seasonality , 0 where seasonality is 0
                if spec['modelling_start_index'] or spec['modelling_end_trim']:
                    deb_by_sea = np.divide(dependent_col_trim, seasonality_col, out=np.zeros_like(seasonality_col), where=seasonality_col!=0)
                    correct_pos = np.zeros_like(dependent_col)
                    correct_pos[spec['modelling_start_index']:spec['modelling_end_trim']] = deb_by_sea
                else:
                    correct_pos = deb_by_sea = np.divide(dependent_col, seasonality_col, out=np.zeros_like(seasonality_col), where=seasonality_col!=0)
                transformed_support[:,pos] *= correct_pos
            else:
                transformed_support[:,pos] = transformed_support[:,pos] * geo_df[postmul[var]]
        elif isinstance(postmul[var], list):
            for v in postmul[var]:
                transformed_support[:,pos] = transformed_support[:,pos] * geo_df[v]
        elif isinstance(postmul[var], float) or isinstance(postmul[var], int):
            transformed_support[:,pos] *= postmul[var]
    
    # 6 Zerofy rows with dependent zero 
    if exclude_dependent_zero:
        mask = dependent_col != 0
        transformed_support = (transformed_support.T * mask).T
        trans_support_no_postmul = (trans_support_no_postmul.T * mask).T  #new added
        support = (support.T * mask).T
        # support_transformation_data = (support_transformation_data.T * mask).T
        # support_nontransformation_data = (support_nontransformation_data.T * mask).T

    # 7 Period Aggregations - dependent, transformed and 2 support
    transformed_support_agg = []
    support_agg = []
    dependent_agg = []
    for p_key, splitter in reporting_period_index.items():
        transformed_support_agg.append(transformed_support[splitter[0]:splitter[1] + 1].sum(axis=0))
        dependent_agg.append(dependent_col[splitter[0]:splitter[1] + 1].sum())
        support_agg.append(support[splitter[0]:splitter[1] + 1].sum(axis=0))

    # 8 Bounds calc
    agg_constraint_cutoff = spec["config"]['agg_constraint_cutoff']
    agg_period_search = spec["config"]['agg_period_search']
    contri_override = spec['contri_override'][file_name] if file_name in spec['contri_override'].keys() else []
    
    #Condition to use User defined bounds or engine calculated
    if spec['is_custom_dep_agg']:
        is_store_need_scaled_bounds = True if spec['stores_need_bound'][file_name] else False
        # print('is_store_need_scaled_bounds: ', file_name, ':', is_store_need_scaled_bounds)
        # print('contribution_range main: ', contribution_range['DEP_ALL'])
        bounds, scale_vector, system_unbound = bounds_calc(contribution_range, transformed_support_agg, support_agg, dependent_agg, allvar_support_colnames, list(reporting_period_index.keys()), agg_constraint_cutoff, agg_period_search, contri_override, spec, is_store_need_scaled_bounds, None)
    elif spec["is_volumetric_constraints"]:
        bounds, scale_vector, system_unbound = bounds_calc(contribution_range, transformed_support_agg, support_agg, dependent_agg, allvar_support_colnames, list(reporting_period_index.keys()), agg_constraint_cutoff, agg_period_search, contri_override, spec, None, spec["dependent_distribution"][file_name])
    else :
        bounds, scale_vector, system_unbound = bounds_calc(contribution_range, transformed_support_agg, support_agg, dependent_agg, allvar_support_colnames, list(reporting_period_index.keys()), agg_constraint_cutoff, agg_period_search, contri_override, spec)
    
    # bounds, scale_vector, system_unbound = bounds_calc(contribution_range, transformed_support_agg, support_agg, dependent_agg, allvar_support_colnames, list(reporting_period_index.keys()), agg_constraint_cutoff, agg_period_search, contri_override)
    # Trim according to modelling period
    if spec['modelling_start_index'] or spec['modelling_end_trim']:
        transformed_support = transformed_support[spec['modelling_start_index']:spec['modelling_end_trim']]
        trans_support_no_postmul = trans_support_no_postmul[spec['modelling_start_index']:spec['modelling_end_trim']] #new added
        dependent_col = dependent_col[spec['modelling_start_index']:spec['modelling_end_trim']]
    # Add Seasonality
    if spec["config"]["include_seasonality"]:
        if exclude_dependent_zero:
            seasonality_col[dependent_col == 0] = 0
        transformed_support = np.c_[transformed_support, seasonality_col]
        trans_support_no_postmul = np.c_[trans_support_no_postmul, seasonality_col] #new added
        transformed_support_colnames.append("Seasonality")
        bounds.append([-np.inf,np.inf])
        scale_vector = np.append(scale_vector, [0])
    # Add Intercept
    if spec["config"]["include_intercept"]:
        if(file_name in spec["geo_ex_intercept"].keys()):
            intercept = np.zeros(len(dependent_col))
            bounds.append([0,0])
        else:
            intercept = np.ones(len(dependent_col))
            bounds.append([-np.inf,np.inf])
        if exclude_dependent_zero:
            intercept[dependent_col == 0] = 0
        transformed_support = np.c_[transformed_support, intercept]
        trans_support_no_postmul = np.c_[trans_support_no_postmul, intercept]
        transformed_support_colnames.append("Intercept")
        scale_vector = np.append(scale_vector, [0])

    return np.round(transformed_support, 8), np.round(trans_support_no_postmul, 8), dependent_col, transformed_support_colnames, bounds, scale_vector, system_unbound, dependent_agg

def gen_seasonality(dependent, distinct_week, exclude_week = [], exclude_dependent_zero=False):
    row_count = len(dependent)
    mod = row_count % distinct_week
    if exclude_week and len(exclude_week) > 0:
        for i in exclude_week:
            dependent[i] = np.nan
    if exclude_dependent_zero:
        dependent[dependent == 0] = np.nan
    if mod > 0 :
        dependent = np.append(dependent, [np.nan] * (distinct_week - mod))
    stream = np.nanmean(np.array(np.split(dependent, len(dependent) / distinct_week)), axis=0)
    stream[np.isnan(stream)] = 0
    return np.array((np.tile(stream, int(row_count / distinct_week) + 1))[:row_count])

@jit(nopython=True)
def scale(array):
    max_val =  np.abs(array).max()
    factor = 1
    if(max_val > 300):
       factor =  300 / max_val
       array = array * factor
    return array, factor

@jit(nopython=True)
def shift(arr, num, fill_value=0): # fastest shifter from stackoverflow
    result = np.empty_like(arr)
    if num > 0:
        result[:num] = fill_value
        result[num:] = arr[:-num]
    elif num < 0:
        result[num:] = fill_value
        result[:num] = arr[-num:]
    else:
        result[:] = arr
    return result

@jit(nopython=True)
def pvalTransformation(var_data, learn, decay, initial_pval = 0):
    val = 1 - ((1 - ( initial_pval  * exp(-decay) )) / ( exp((var_data[0] * learn)/100)))
    var_data[0] = val if val != 'nan' else 0
    for i in range(1, len(var_data)):
        val = 1 - ((1 - ( var_data[i-1]  * exp(-decay) )) / ( exp((var_data[i] * learn)/100)))
        var_data[i] = val if val != 'nan' else 0
    return var_data

# @jit(nopython=True)
def aplTransformation(var_data, adstock, power = 1, lag = 0):
    result = np.copy(var_data)
    initial_apl = 0
    result[0] = var_data[0] + (adstock * initial_apl)
    for i in range(1, len(var_data)):
        result[i] = result[i] + (adstock * result[i-1])
    # Negative numer to the power doesnt work in loop - https://stackoverflow.com/questions/49029019/in-python-3-6-why-does-a-negative-number-to-the-power-of-a-fraction-return-nan
    sign = (result < 0).astype(int) + 1
    sign[sign == 2] = -1
    result = (np.abs(result) ** power) * sign
    result = shift(result, lag)
    return result

# @jit(nopython=True)
def raTransformation(var_data, window = 5): # only central, This will have a complexity of only O(N)
    l = len(var_data)
    result = np.empty(l, dtype=float)
    half = math.floor(window / 2)
    for i in range(half):
        #print(i, " 1 - ", var_data[0 : half+i+1])
        result[i] =  np.mean(var_data[0 : half + i + 1])
    for i in range(half, l - half):
        #print(i, " 2 - ", var_data[i - half: i + half + 1])
        result[i] =  np.mean(var_data[i - half: i + half + 1])
    for i in range(l - half, l):
        #print(i, " 3 - ", var_data[i - half : l])
        result[i] =  np.mean(var_data[i - half : l])
    return result

def select_variables(corr_sum, all_trans_names, transformation_sep):
    selected_variables = {}
    var_name = transformation_sep[0]
    var_max = -np.Inf
    selected_variables[var_name] = all_trans_names[0]
    for i in range(len(transformation_sep)):
        if(var_name == transformation_sep[i]): # comparison of same var
            if corr_sum[i] > var_max:
                selected_variables[var_name] = all_trans_names[i]
                var_max = corr_sum[i]
        else:
            var_max = corr_sum[i]
            var_name = transformation_sep[i]
            selected_variables[var_name] = all_trans_names[i]
    return selected_variables

def run_correlation(all_trans, dependent_col, same_check, transformation_sep, debug=False):
    debug_data = False
    all_trans_names = list(all_trans.keys())
    if len(all_trans_names) == 0:
        return {}, {}
    # calculate coor including dependent and then removing it.
    all_trans['dependent'] = dependent_col
    corr = np.array(list(all_trans.values()))
    valid = ~np.isnan(corr).any(axis=0) #https://stackoverflow.com/questions/35933201/why-is-pandas-and-numpy-producing-different-results-for-pairwise-correlations-wi
    numpy_corr = np.corrcoef(corr[:, valid])
    numpy_corr = np.nan_to_num(numpy_corr)
    corr_with_dependent = numpy_corr[-1][0:-1] # last row with last element removed
    del all_trans['dependent']
    numpy_corr = np.delete(np.delete(numpy_corr, -1, 0), -1, 1) # delete dependent from row and col
    # This is to remove coorelation with same variables, same_check is generated once for performance
    numpy_corr = numpy_corr * same_check
    # Corr Algo
    numpy_corr[numpy_corr < 0] = 0
    corr_inner_sum = numpy_corr.sum(axis=0) + numpy_corr.sum(axis=1)
    corr_count = np.count_nonzero(numpy_corr, axis=0) + np.count_nonzero(numpy_corr, axis=1)
    corr_inner = np.divide(corr_inner_sum, corr_count, out=np.zeros_like(corr_inner_sum), where=corr_count != 0)
    corr_sum = (1 - corr_inner) + corr_with_dependent
    # pd.DataFrame([1 - corr_inner_mean, corr_with_dependent, corr_sum], index=["inner", "dep", "sum"], columns=all_trans_names).T.to_csv("CORR-TEST4.csv")
    selected_variables = select_variables(corr_sum, all_trans_names, transformation_sep)
    if debug:
        debug_data = pd.DataFrame([corr_inner, corr_with_dependent, corr_sum], index=["inner", "dep", "sum"], columns=all_trans_names)[list(selected_variables.values())].T
    return selected_variables, debug_data # Dict [Variable : variable__Transformation]

def run_transformations(geo_df, transformed_colnames, initial_pval, LD):
    all_trans = {}
    scale_factor = {}
    for trans in transformed_colnames:
        variable, all_transformation = trans.split("__")
        transformations = all_transformation.split("#")
        var_data = geo_df[variable]
        trans_data = np.copy(var_data)
        for transformation in transformations:
            _split = transformation.split("_")
            trans_method = _split[0]
            vals = _split[1].split("|")
            if(trans_method == "PVAL"):
                if variable in scale_factor.keys() :
                    if(scale_factor[variable] != 1): trans_data = scale_factor[variable] * trans_data
                else :
                    trans_data, factor = scale(trans_data) # scale below 300
                    scale_factor[variable] = factor
                initial = initial_pval[variable] if variable in initial_pval.keys() else 0
                trans_data = pvalTransformation(trans_data, LD[int(vals[0])], LD[int(vals[1])], float(initial))
            elif(trans_method == "APL"):
                trans_data = aplTransformation(trans_data, float(vals[0]), float(vals[1]), int(vals[2]))
            elif(trans_method == "RA"):
                trans_data = raTransformation(trans_data, int(vals[0]))
            elif(trans_method == "LAG"):
                trans_data = shift(trans_data, int(vals[0]))
        all_trans[trans] = trans_data
    return all_trans, scale_factor

@jit(nopython=True) # 1.69 --> 0.03 sec run 10,000
def calc_r2_dw_mape(dependent, predicted):
    actualSq = dependent ** 2
    predictedSq = predicted ** 2
    actualPredicted = dependent * predicted
    residual = dependent - predicted
    residualSq = residual ** 2
    residualLag = shift(residual, 1)
    resSq = (residualLag - residual) ** 2
    resSq[0] = 0
    # Sums
    count = len(dependent)
    sumActual = dependent.sum()
    sumPredicted = predicted.sum()
    sumActualSq = actualSq.sum()
    sumPredictedSq = predictedSq.sum()
    sumActualPredicted = actualPredicted.sum()
    sumResidualSq = residualSq.sum()
    # R2
    def1 = (count * sumPredictedSq) - (sumPredicted ** 2)
    def2 = (count * sumActualSq) - (sumActual ** 2)
    den = (abs(def1*def2)) ** .5
    num = (count * sumActualPredicted) - (sumActual*sumPredicted)
    r2 = (np.divide(num,den) ** 2) * 100
    # DW
    dw = np.divide(resSq.sum(),sumResidualSq)
    # Mape
    pct_err = np.empty(count)
    #dependentMean = np.mean(dependent) -> This might be a better calc.
    for j in range(count):
        if dependent[j] != 0:
            pct_err[j] = (dependent[j] - predicted[j]) / dependent[j]
        else:
            pct_err[j] = 0 #predicted[j] / dependentMean
    mape = np.mean(np.abs(pct_err)) * 100
    return r2,dw,mape,sumResidualSq, sumActual

def stats_calc(dependent, predicted, transformed_support, coeff, sumResidualSq, intercept = True, store_count = 1): #Main Function , PVAL is not done here, to enhance performce -> this can now be nopython
    shape = transformed_support.shape
    if shape[0] < shape[1]:
        # arr_add = np.mean(transformed_support, axis=0)
        # zeroes_add = np.repeat([arr_add], shape[1] - shape[0], axis=0)
        # transformed_support_extended = np.append(transformed_support, zeroes_add, axis=0)
        transformed_support_extended = transformed_support
        arr_avg = np.mean(transformed_support, axis=0)
        arr_min = arr_avg * 0.8
        arr_max = arr_avg * 1.2
        idx = np.arange(1,shape[1] - shape[0] + 1)
        for i in idx:
            df = pd.DataFrame(arr_min, columns=["arr_min"])
            df["arr_max"] = arr_max
            df['rand_between'] = df.apply(lambda x: np.random.uniform(x.arr_min, x.arr_max), axis=1)
            npar = df["rand_between"].to_numpy()
            transformed_support_extended = np.vstack([transformed_support_extended, npar])
        u,d,vt = np.linalg.svd(transformed_support_extended,full_matrices=True) # Faster than inversion of matrix - https://stats.stackexchange.com/questions/6731/using-singular-value-decomposition-to-compute-variance-covariance-matrix-from-li
    else:
        u,d,vt = np.linalg.svd(transformed_support,full_matrices=True)
    # d = (d ** -2) * np.eye(len(d))
    d = (d ** -2) #* np.eye(len(d))
    d[d == np.inf] = 0
    d = d*np.eye(len(d))
    tt = np.dot(vt.T, d)
    var_cov = (tt * vt.T).sum(-1) # var_cov =  np.diag(np.dot(tt, vt)) # https://stackoverflow.com/questions/14758283/is-there-a-numpy-scipy-dot-product-calculating-only-the-diagonal-entries-of-the
    transformed_support_sum = transformed_support.sum(axis = 0)
    transformed_support_mask = transformed_support_sum == 0
    var_cov[transformed_support_mask] = 0

    if intercept:
        transformed_support = transformed_support[:, :-1] # remove intercept

    row_count = shape[0]
    var_count = shape[1] - np.sum(transformed_support_mask)
    residualDF = (row_count*store_count) - var_count - 1
    if residualDF < 1: residualDF = 1

    residual_SS = sumResidualSq
    std_error_ss = (residual_SS / residualDF)
    std_error_col_ss = (std_error_ss * var_cov)
    # For VIF
    support_dev = (transformed_support - transformed_support.mean(axis = 0)) ** 2
    support_dev_sum = support_dev.sum(axis = 0)

    if intercept:
        VIF = support_dev_sum * (std_error_col_ss[0:-1]) / (residual_SS / residualDF)
        VIF[VIF < 1] = np.nan
        VIF = np.append(VIF, np.nan)
    else:
        VIF = support_dev_sum * (std_error_col_ss) / (residual_SS / residualDF)
        VIF[VIF < 1] = np.nan

    std_error_col = std_error_col_ss ** .5
    tstat = coeff  / std_error_col
    factor = std_error_col * 1.9617999999998 # hard code to 95
    lower = coeff - factor
    upper = coeff + factor

    stats_dict = {
        'Coefficient': coeff,
        'Std. Error': std_error_col,
        'T Stat': tstat,
        'Pval': [],
        'Lower 95%': lower,
        'Upper 95%': upper,
        'VIF': VIF
    }
    return stats_dict, residual_SS, support_dev_sum, std_error_col_ss, transformed_support_mask.astype(int)

def overall_stats_gen(sum_coeff, sum_support_dev_sum, sum_std_error_col_ss, sum_residual_SS, sum_transformed_support_mask, no_of_stores, row_count, intercept, sum_support_optimized, sumResidualSq):

    var_count = len(sum_support_dev_sum) - np.sum(sum_transformed_support_mask[sum_transformed_support_mask == no_of_stores])# handles intercept automatically
    residualDF = (row_count * no_of_stores) - var_count - 1
    if residualDF < 1: residualDF = 1
    corrected_no_of_stores = no_of_stores - sum_transformed_support_mask #

    std_error_ss = sum_residual_SS / residualDF
    coeff = sum_coeff / no_of_stores # Avg
    std_error_col_ss = np.divide(sum_std_error_col_ss, corrected_no_of_stores, out=np.zeros_like(corrected_no_of_stores), where=corrected_no_of_stores != 0) # Avg
    support_dev_sum = sum_support_dev_sum / no_of_stores # Avg

    if intercept:
        VIF = support_dev_sum * (std_error_col_ss[0:-1]) / std_error_ss
        print("VIF =", VIF.sum())
        VIF[VIF < 1] = np.nan
        VIF = np.append(VIF, np.nan)
    else:
        VIF = support_dev_sum * (std_error_col_ss) / std_error_ss
        VIF[VIF < 1] = np.nan

    std_error_col = std_error_col_ss ** .5
    tstat = coeff  / std_error_col
    factor = std_error_col * 1.9617999999998 # hard code to 95
    lower = coeff - factor
    upper = coeff + factor

    #=======std err calc new : stats_dict from total result====================
    # transformed_support = sum_support_optimized/coeff
    transformed_support = np.divide(sum_support_optimized, coeff, out=np.zeros_like(sum_support_optimized), where=coeff!=0) #sum_support_optimized/coeff
    #https://stackoverflow.com/questions/26248654/how-to-return-0-with-divide-by-zero
    stats_dict, residual_SS, support_dev_sum, std_error_col_ss, transformed_support_mask = stats_calc(0, 0, transformed_support, coeff, sumResidualSq, intercept, no_of_stores)

    #--------------------------------------------

    # stats_dict = {
    #     'Coefficient': coeff,
    #     'Std. Error': std_error_col,
    #     'T Stat': tstat,
    #     'Pval': [],
    #     'Lower 95%': lower,
    #     'Upper 95%': upper,
    #     'VIF': VIF
    # }
    total_transformed_support = transformed_support
    return stats_dict, total_transformed_support 

@jit(nopython=True)
def objectiveFunction(x0, d, a):
    p = (x0 * d).sum(axis=1)
    r = a - p
    e = r ** 2
    g = (d.T * -2 * r).T
    return e.sum(), g.sum(axis=0)

@jit(nopython=True) # 1.69 --> 0.03 sec run 10,000
def calc_r2_complex(dependent, predicted):
    actualSq = dependent ** 2
    predictedSq = predicted ** 2
    actualPredicted = dependent * predicted
    # Sums
    count = len(dependent)
    sumActual = dependent.sum()
    sumPredicted = predicted.sum()
    sumActualSq = actualSq.sum()
    sumPredictedSq = predictedSq.sum()
    sumActualPredicted = actualPredicted.sum()
    # R2
    def1 = (count * sumPredictedSq) - (sumPredicted ** 2)
    def2 = (count * sumActualSq) - (sumActual ** 2)
    den = (abs(def1*def2)) ** .5
    num = (count * sumActualPredicted) - (sumActual*sumPredicted)
    r2 = (np.divide(num,den) ** 2) * 100
    return r2

def binary_search(low, high, x, residual, distortion, predicted, dependent_col, last_val):
    allow = 0.0001
    if high >= low:
        mid = (high + low)  * 0.5
        stocastic = ( mid * distortion) + residual
        newPred = predicted + stocastic
        value = calc_r2_complex(dependent_col, newPred)
        # print(mid, value)
        if last_val == value:
            return mid
        elif mid > 0.999:
            return 1
        elif value > x * (1 - allow) and value < x * (1 + allow):
            return mid
        elif value < x:
            return binary_search(low, mid, x, residual, distortion, predicted, dependent_col, value)
        else:
            return binary_search(mid, high, x, residual, distortion, predicted, dependent_col, value)
    else:
        return -1


def gen_stochastic_sea(dependent_col, predicted, spec, gof):
    residual = dependent_col - predicted
    rolling_avg = raTransformation(residual, gof)
    return rolling_avg

def scale_opt(support, bounds, scale):
    bounds = (bounds.T / scale).T
    support = support * scale
    return support, bounds

def run_optimization_fast(support, dependent_col, all_cols, bounds=None, scale_vector=None):
    method = "TNC"
    ones = np.zeros(support.shape[1])
    res = minimize(objectiveFunction, ones, jac=True, args=(support, dependent_col), method="TNC", bounds=bounds, options={"maxiter": support.shape[1]*10})
    if res.status == 6: # "Unable to progress"
        tnc_res = res
        res = minimize(objectiveFunction, ones, args=(support, dependent_col), method="L-BFGS-B", bounds=bounds, jac=True, options={"maxiter": support.shape[1]*200})
        if res.fun < tnc_res.fun:
            method = "L-BFGS-B"
        else:
            res = tnc_res
    res.success = False
    return res.x, res.success, res.message, method

def run_optimization(support, dependent_col, all_cols, bounds=None, scale_vector=None, dynamic_opt_itr = 200):
    # return run_optimization_exp(support, dependent_col, all_cols, bounds=None, scale_vector=None)
    dynamic_opt_itr = int(dynamic_opt_itr)
    method = "TNC SCALE"
    bounds = np.array(bounds)
    ones = np.zeros(support.shape[1])
    b_min = scale_vector.min()
    b_max = scale_vector.max()
    lowest_error = np.inf
    final_scale = None
    best_res = None
    selected_i = None
    # print(bounds)
    for i in [0.1,1,2,3,5,8,10]: # Try 6 different scales when TNC failes to progress or linear search failes - status 4 and 6.
        scale = np.interp(scale_vector, (b_min, b_max), (1, 10 * i))
        scale_support, scale_bounds = scale_opt(support, bounds, scale)
        res = minimize(objectiveFunction, ones, jac=True, args=(scale_support, dependent_col), method="TNC", bounds=scale_bounds, options={"maxiter": support.shape[1]*dynamic_opt_itr})
        if res.fun < lowest_error:
            lowest_error = res.fun
            best_res = res
            final_scale = scale
            selected_i = i
        if not (res.status == 6 or res.status == 4):
            break

    best_res.x = best_res.x * final_scale
    if not best_res.success: # Hopefully Code Never Comes Here
        res = minimize(objectiveFunction, ones, jac=True, args=(support, dependent_col), method="L-BFGS-B", bounds=bounds, options={"maxiter": support.shape[1]*dynamic_opt_itr, 'ftol': 1e-15, 'maxfun': support.shape[1]*dynamic_opt_itr})
        if res.fun < lowest_error:
            lowest_error = res.fun
            best_res = res
            final_scale = 1
            method = "L-BFGS-B + TNC SCALE"

    # For Printing
    best_res.success = False
    method = method + " | " + str(selected_i) + " | i = " + str(i) + " | opt_itr = " + str(dynamic_opt_itr)
    return best_res.x, best_res.success, best_res.message, method

def run_optimization_exp(support, dependent_col, all_cols, bounds=None, scale_vector=None):
    method = "TNC SCALE"
    bounds = np.array(bounds)
    ones = np.ones(support.shape[1])
    b_min = scale_vector.min()
    b_max = scale_vector.max()
    lowest_error = np.inf
    final_scale = None
    best_res = None

    selected_i = None
    for i in [1,2,3,4,5,6,7,8,9,10]: # Try 10 different scales when TNC failes to progress or linear search fails - status 4 and 6.
        scale = np.interp(scale_vector, (b_min, b_max), (1, 10 * i))
        scale_support, scale_bounds = scale_opt(support, bounds, scale)
        res = minimize(objectiveFunction, ones, jac=True, args=(scale_support, dependent_col), method="TNC", bounds=scale_bounds, options={"maxiter": support.shape[1]*200})
        if res.fun < lowest_error:
            lowest_error = res.fun
            best_res = res
            final_scale = scale
            selected_i = i
        if not (res.status == 6 or res.status == 4):
            break

    best_res.x = best_res.x * final_scale
    if not best_res.success: # Hopefully Code Never Comes Here
        res = minimize(objectiveFunction, ones, jac=True, args=(support, dependent_col), method="L-BFGS-B", bounds=bounds, options={"maxiter": support.shape[1]*200, 'ftol': 1e-15, 'maxfun': support.shape[1]*200})
        if res.fun < lowest_error:
            lowest_error = res.fun
            best_res = res
            final_scale = 1
            method = "L-BFGS-B + TNC SCALE"

    # For Printing
    best_res.success = False
    method = method + " | " + str(selected_i) + " | i = " + str(i)
    return best_res.x, best_res.success, best_res.message, method

def files_to_core_split(files, cpu):
    splitBy = int(len(files) / cpu)
    splitList = [files[i:i + splitBy] for i in range(0, len(files), splitBy)]
    if len(splitList) > cpu:
        extra = splitList.pop()
        for i in range(len(extra)): # distribute extra from begining
            splitList[i].append(extra[i])
    return splitList

def periodFinder(period_s, period):
    for index, period_s_date in period_s.iteritems():
        if period_s_date < period:
            continue
        else:
            return index
        
def combine_granular(input_path, output_path, filename, is_dma = False, rand_map_path = '', remove_from_name = "__result", zip_also=False):
    if os.path.exists(output_path + "\\" + filename + ".csv"):
        os.remove(output_path + "\\" + filename + ".csv")
    files = [f for f in os.listdir(input_path) if isfile(join(input_path, f))]
    # print("input_path: ", input_path)
    # print("output_path: ", output_path)
    # print("filename: ", filename)
    # print("is_dma: ", is_dma)
    # Result file's header:
    f = open(os.path.join(input_path, '../result.csv'))
    headers = f.readline()
    locOfComma = [i for i, n in enumerate(headers) if n == ',']
    headers = headers[:locOfComma[0]] + headers[locOfComma[1]:] # remove Geography
    
    if is_dma:
        dma_dict = dict()
        period_s = pd.read_csv(os.path.join(input_path, '../result.csv'))["Period"]

        final_headers = [f.strip() for f in headers.split(',')]
        final_headers.insert(0,"DMA")
        gran_only_headers = [f.split('__')[0] for f in headers.split(',') if f != "Period"]
        newDF_to_sum = pd.DataFrame(columns =  gran_only_headers)

        gran_only_headers_append = [f.split('__')[0] for f in headers.split(',')]
        gran_only_headers_append.insert(0,"DMA")

        newDF_to_append = pd.DataFrame(columns =  gran_only_headers_append)
        dma_map = pd.read_csv(os.path.join(input_path, '../../../../generic/DMA_Store_Map.csv'))

        files_without_result = [f.split('__result.csv')[0] for f in files]
        dma = [f[1] for f in dma_map.to_numpy()]
        # print("files_without_result: ", files_without_result)
        # print("dma: ", dma)
        matched_geo = np.intersect1d(files_without_result, dma)
        # print("matched_geo: ", matched_geo)
        if len(matched_geo) > 0:
            for i in dma_map.to_numpy():
                if i[0] not in dma_dict:
                    dma_dict[i[0]] = []
                    dma_dict[i[0]].append(i[1])
                else:
                    dma_dict[i[0]].append(i[1])  
            for dma in dma_dict.keys():
                for geo in dma_dict[dma]:
                    _file = geo+"__result.csv"
                    if _file in files:
                        data_d = pd.read_csv("\\".join([input_path, _file]))
                        data_d_no_Period = data_d.drop(['Period'], axis=1)
                        data_d_no_Period.columns = gran_only_headers
                        newDF_to_sum = newDF_to_sum.add(data_d_no_Period, fill_value=0)
                newDF_to_sum.insert(loc=0, column='Period', value = period_s)
                newDF_to_sum.insert(loc=0, column="DMA", value = [dma for f in period_s])

                newDF_to_append = newDF_to_append.append(newDF_to_sum, ignore_index=True)
                newDF_to_sum =  pd.DataFrame(columns = gran_only_headers) # initializing again
            newDF_to_append.columns = final_headers
            newDF_to_append.to_csv(output_path + '\\' + filename + '.csv' , index=False) 
        # print("Done")    
    elif rand_map_path:
        fr = open(os.path.join(rand_map_path))
        map_header =  fr.readline().strip().split(',')
        map_header = map_header[::-1]
        data = pd.read_csv(rand_map_path, low_memory=False, memory_map=True, engine='c')[map_header]
        rand_mapping_dict = data.set_index(map_header[-1]).T.to_dict('list')  
        fout=open(output_path + "\\" + filename + ".csv","w")
        fout.write(",".join(map_header) + "," + headers)
        f.close() # not really needed
        for _file in files:
            f = open(input_path + "\\" + _file)
            next(f) # skip the header
            for line in f:
                fout.write(",".join(rand_mapping_dict[_file.split(remove_from_name)[0]]) + "," +_file.split(remove_from_name)[0] + "," + line)
            f.close() # not really needed
        fout.close()
    else:
        fout=open(output_path + "\\" + filename + ".csv","w")
        fout.write("Geography," + headers)
        f.close() # not really needed
        for _file in files:
            f = open(input_path + "\\" + _file)
            next(f) # skip the header
            for line in f:
                fout.write(_file.split(remove_from_name)[0] + "," + line)
            f.close() # not really needed
        fout.close()
    if zip_also:
        archive = py7zr.SevenZipFile(output_path + "\\" + filename + ".7z", 'w')
        archive.writeall(output_path + "\\" + filename + ".csv", filename + ".csv")
        archive.close()

def combine_randomize_results(input_path, output_path, filename, level = '', remove_from_name = "__result", zip_also=False):
    # input_path = input_path + "\\" + level + "_result"
    
    files = [f for f in os.listdir(input_path) if isfile(join(input_path, f))]
    f = open(os.path.join(input_path, '../../result.csv'))
    headers = f.readline()
    locOfComma = [i for i, n in enumerate(headers) if n == ',']
    headers = headers[:locOfComma[0]] + headers[locOfComma[1]:] # remove Geography
    
    fout=open(output_path + "\\" + filename + ".csv","w")
    fout.write(level + "," + headers)
    f.close() # not really needed
    for _file in files:
        f = open(input_path + "\\" + _file)
        next(f) # skip the header
        for line in f:
            fout.write(_file.split(remove_from_name)[0] + "," + line)
        f.close() # not really needed
    fout.close()
    if zip_also:
        archive = py7zr.SevenZipFile(output_path + "\\" + filename + ".7z", 'w')
        archive.writeall(output_path + "\\" + filename + ".csv", filename + ".csv")
        archive.close()

def zipGranular(output_path, input_path, foldername):
    archive = py7zr.SevenZipFile(output_path, 'w')
    archive.writeall(input_path, foldername)
    archive.close()
    return

def rapid_refresh(old_path, new_path, output_path):
  #  from os import path
    status_path = output_path + "\\status.txt"
    g_old = old_path + "\\granular_result"
    g_new = new_path + "\\granular_result"
    g_old_files = [f for f in os.listdir(g_old) if os.path.isfile(os.path.join(g_old, f))]
    g_new_files = [f for f in os.listdir(g_new) if os.path.isfile(os.path.join(g_new, f))]
    diff = list(set(g_old_files)-set(g_new_files))
    if len(diff) > 0:
        communicator("There is a difference in geography files - " + str(diff), status_path)
        raise("There is a difference in geography files - " + str(diff))
    # Operations with 1st file
    df_old = pd.read_csv(g_old + "\\" + g_old_files[0], low_memory=False, memory_map=True, engine='c')
    df_new = pd.read_csv(g_new + "\\" + g_new_files[0], low_memory=False, memory_map=True, engine='c')
    col_old = df_old.columns
    col_new = df_new.columns
    col_old = [i.split("__")[0] for i in col_old]
    col_new = [i.split("__")[0] for i in col_new]
    var_diff = (list(list(set(col_old)-set(col_new)) + list(set(col_new)-set(col_old))))
    if (len(var_diff) > 0):
        communicator("There is a difference variable list - " + str(var_diff), status_path)
        raise("There is a difference variable list - " + str(var_diff))
    period_old = df_old.iloc[:, 0]
    period_new = df_new.iloc[:, 0]
    if(len(period_old) > len(period_new)):
        communicator("New refresh model has lesser time period", status_path)
        raise("New refresh model has lesser time period")
    for i in range(len(period_old)):
        if period_old[i] != period_new[i]:
            communicator("Period did not match at location - " + str(i), status_path)
            raise("Period did not match at location - " + str(i))
    newdata_start_index = i + 1
    try:
        os.mkdir(output_path + '/stats')
    except:
        pass
    try:
        os.mkdir(output_path + '/granular_result')
    except:
        pass
    # Create Granulars
    level_stats = {}
    for _file in g_new_files:
        if _file in g_old_files:
            df_old = pd.read_csv(g_old + "\\" + _file, low_memory=False, memory_map=True, engine='c')
            df_new_only = pd.read_csv(g_new + "\\" + _file, skiprows=newdata_start_index, low_memory=False, memory_map=True, engine='c')
            df_new_only.columns = col_new
            df_old.columns = col_old
            df_final = pd.concat([df_old, df_new_only])
        else:
            df_final = df_old = pd.read_csv(g_new + "\\" + _file, low_memory=False, memory_map=True, engine='c')
        df_final.to_csv(output_path + '/granular_result/' + _file, index=False)
        predicted = df_final["Predicted"].to_numpy()
        dependent = df_final.iloc[:,-1].to_numpy()
        r2, dw, mape, sumResidualSq, dep_agg = calc_r2_dw_mape(dependent, predicted)
        level_stats[_file.split('__')[0]] = [r2, dw, mape, dep_agg]

    # Create Main result
    df_old = pd.read_csv(old_path + "\\result.csv", low_memory=False, memory_map=True, engine='c')
    col_new = pd.read_csv(new_path + "\\result.csv", nrows=1, low_memory=False, memory_map=True, engine='c').columns
    df_new_only = pd.read_csv(new_path + "\\result.csv", skiprows=newdata_start_index, low_memory=False, memory_map=True, engine='c')
    col_old = df_old.columns
    col_old = [i.split("__")[0] for i in col_old]
    col_new = [i.split("__")[0] for i in col_new]
    df_new_only.columns = col_new
    df_old.columns = col_old
    df_final = pd.concat([df_old, df_new_only])
    df_final.to_csv(output_path + '\\result.csv', index=False)
    # Create RDM
    predicted = df_final["Predicted"].to_numpy()
    dependent = df_final.iloc[:,-1].to_numpy()
    r2, dw, mape, sumResidualSq, dep_agg = calc_r2_dw_mape(dependent, predicted)
    r2_dw_mape = pd.DataFrame([[r2,dw,mape]], columns=["R-Square", "DW", "MAPE"])
    r2_dw_mape.to_csv(output_path + '\\RDM.csv',  index=False)
    print(r2_dw_mape)
    # Level Stats
    level_stats = pd.DataFrame(level_stats, index=['R - Square', 'DW', 'MAPE',"Aggregate"]).T
    no_of_stores = len(level_stats[level_stats["DW"] != -1])
    if no_of_stores > 1:
        level_stats["Rank"] = (level_stats["Aggregate"].rank(method="min") - 1) * (100 / no_of_stores) # percentile 0 to 100
    else:
        level_stats["Rank"] = 100
    level_stats["Ratio"] = level_stats["Aggregate"] / level_stats["Aggregate"].sum() * 100
    level_stats.insert(0, "Geography", level_stats.index)
    level_stats.to_csv(output_path + '/level_stats.csv',  index=False)
    return "Success"

def add_sea_and_intercept(split_path, spec, output_path):
    old_path = spec['old_model_path']
    status_path = output_path + "\\status.txt"
    
    g_old = old_path + "\\granular_result"
    g_old_files = [f.split("__result")[0] + ".csv" for f in os.listdir(g_old) if os.path.isfile(os.path.join(g_old, f))]
    
    example_df = pd.read_csv(g_old + "\\" + g_old_files[0].split(".csv")[0] + "__result.csv", nrows=(0))
    usecols = []
    contribution_range = spec['contribution_range']
    if "Seasonality" in example_df.columns: 
        usecols.append("Seasonality")
        contribution_range["Seasonality"] = {'contri_period': 'FIXED', 'min': np.nan, 'max': np.nan, 'sign': np.nan}
        spec['nontransformation_cols'] = spec['nontransformation_cols'] + ["Seasonality"]
        spec['all_cols_read'] = spec['all_cols_read'] + ["Seasonality"]
        spec['var_count'] = spec['var_count'] + 1
    if "Intercept" in example_df.columns: 
        usecols.append("Intercept")
        contribution_range["Intercept"] = {'contri_period': 'FIXED', 'min': np.nan, 'max': np.nan, 'sign': np.nan}
        spec['nontransformation_cols'] = spec['nontransformation_cols'] + ["Intercept"]
        spec['all_cols_read'] = spec['all_cols_read'] + ["Intercept"]
        spec['var_count'] = spec['var_count'] + 1
    if len(usecols) == 0: return spec
    
    spec['contribution_range'] = contribution_range
    spec["config"]["include_seasonality"] = False
    spec["config"]["include_intercept"] = False

    split_files = [f for f in os.listdir(split_path) if os.path.isfile(os.path.join(split_path, f))]
    diff = (list(list(set(g_old_files)-set(split_files)) + list(set(split_files)-set(g_old_files))))
    # These diff can be files which are not generated because dep is 0.
    final_files = list(set(g_old_files) - set(diff))
    for _file in final_files:
        old_geo_df = pd.read_csv(g_old + "\\" + _file.split(".csv")[0] + "__result.csv", usecols=usecols, nrows=(52))
        geo_df = pd.read_csv(split_path + "\\" + _file)
        if "Seasonality" in usecols:
            seasonality = old_geo_df["Seasonality"].to_numpy()
            geo_df["Seasonality"] = np.tile(seasonality, int(len(geo_df) / 52) + 1)[:len(geo_df)]
        if "Intercept" in usecols:
            intercept = old_geo_df["Intercept"][0]
            geo_df["Intercept"] = intercept
        geo_df.to_csv(split_path + "\\" + _file, index=False)
    
    for _file in diff:
        geo_df = pd.read_csv(split_path + "\\" + _file)
        if "Seasonality" in usecols:
            dependent_col = np.array(geo_df[spec['dependent_name']])
            geo_df["Seasonality"] = gen_seasonality(dependent_col, spec['config']['seasonality_cycle'], spec['config']['seasonality_exclude'])
        if "Intercept" in usecols: geo_df["Intercept"] = 1
        geo_df.to_csv(split_path + "\\" + _file, index=False)

    return spec, diff
def combine_stats(input_path, output_path, filename, rand_map_path = '', remove_from_name = "__stats", zip_also=True):
    files = [f for f in os.listdir(input_path) if isfile(join(input_path, f))]
    fout=open(output_path + "\\" + filename + ".csv","w")
    # Result file's header:
    f = open(os.path.join(input_path, '../result_stats.csv'))
    headers = f.readline()
    locOfComma = [i for i, n in enumerate(headers) if n == ',']
    fout.write("Geography," + headers[:-1] + ',Scale\n')
    f.close() # not really needed
    if rand_map_path:
        fr = open(os.path.join(rand_map_path))
        map_header =  fr.readline().strip().split(',')
        map_header = map_header[::-1]
        data = pd.read_csv(rand_map_path, low_memory=False, memory_map=True, engine='c')[map_header]
        rand_mapping_dict = data.set_index(map_header[-1]).T.to_dict('list')  
        fout=open(output_path + "\\" + filename + ".csv","w")
        fout.write(",".join(map_header) + "," + headers)
        f.close() # not really needed
        for _file in files:
            f = open(input_path + "\\" + _file)
            next(f) # skip the header
            for line in f:
                fout.write(",".join(rand_mapping_dict[_file.split(remove_from_name)[0]]) + "," +_file.split(remove_from_name)[0] + "," + line)
            f.close() # not really needed
        fout.close()
    else:
        for _file in files:
            f = open(input_path + "\\" + _file)
            next(f) # skip the header
            for line in f:
                fout.write(_file.split(remove_from_name)[0] + "," + line)
            f.close() # not really needed
        fout.close()
    if zip_also:
        archive = py7zr.SevenZipFile(output_path + "\\" + filename + ".7z", 'w')
        archive.writeall(output_path + "\\" + filename + ".csv", filename + ".csv")
        archive.close()
def combine_randomize_stats(input_path, output_path, filename, level, remove_from_name = "__stats", zip_also=False):
    # input_path = input_path + "\\" + level + "_stats"
    files = [f for f in os.listdir(input_path) if isfile(join(input_path, f))]
    fout=open(output_path + "\\" + filename + ".csv","w")
    # Result file's header:
    f = open(os.path.join(input_path, '../../result_stats.csv'))
    headers = f.readline()
    locOfComma = [i for i, n in enumerate(headers) if n == ',']
    fout.write(level+ "," + headers[:-1] + ',Scale\n')
    f.close() # not really needed
    for _file in files:
        f = open(input_path + "\\" + _file)
        next(f) # skip the header
        for line in f:
            fout.write(_file.split(remove_from_name)[0] + "," + line)
        f.close() # not really needed
    fout.close()
    if zip_also:
        archive = py7zr.SevenZipFile(output_path + "\\" + filename + ".7z", 'w')
        archive.writeall(output_path + "\\" + filename + ".csv", filename + ".csv")
        archive.close()

def core_xlsx_to_json(xlsx_path):
    xlsx_path = "\\".join([path_to_core, xlsx_path])
    business_constraints_sheet = pd.read_excel(xlsx_path, sheet_name='Business Constraints')
    base_settings_sheet = pd.read_excel(xlsx_path, sheet_name='Base Settings')
    config_sheet = pd.read_excel(xlsx_path, sheet_name='Config')

    # Read Business Constraints Sheet
    business_constraints_sheet.columns = business_constraints_sheet.columns.str.upper().str.strip()
    constraints_columns = {
        'INCLUDE':'include',
        'TYPE':'type',
        'VARIABLE':'variable',
        'VARIABLE DESCRIPTION':'desp',
        'PERIOD':"contri_period",
        'MIN (%)':"min",
        'MAX (%)':'max',
        'SIGN':'sign',
        'TRANSFORMATION':'transformation',
        'TRANS. VALUES':'transformation_values',
        'POST MULTIPLICATION':'postmul',
        'DEPENDENT':'dependent'
    }
    cols_for_strip = ["TYPE","TRANSFORMATION","SIGN"]
    business_constraints_sheet[cols_for_strip] = business_constraints_sheet[cols_for_strip].apply(lambda x: x.str.strip())
    business_constraints_sheet = business_constraints_sheet.rename(columns = constraints_columns)
    business_constraints_sheet['include'] = business_constraints_sheet['include'].map({"YES":"1" ,"NO":"0"})

    business_constraints_array = business_constraints_sheet.to_dict(orient='records')

    # Read Business Config Sheet
    # Pval config
    pval_config_sheet = config_sheet.iloc[1:,0:2] # removing first two columns containg headers
    pval_config_sheet.columns = ['definition', 'value']
    config_json = dict(zip(pval_config_sheet["definition"], pval_config_sheet["value"]))

    # Seasonality Exclude Dates
    seasonality_exclude = config_sheet.iloc[:,2:3]
    seasonality_exclude = seasonality_exclude.dropna()
    seasonality_exclude = seasonality_exclude.astype(int)
    seasonality_exclude.columns = ['value']
    seasonality_exclude_json = {"seasonality_exclude":seasonality_exclude["value"].tolist()}

    # Advanced Option
    advanced_options = config_sheet.iloc[:,4:6]
    advanced_options.columns = ['key','value']    
    advanced_options = advanced_options.dropna()
    advanced_options['value'] = advanced_options['value'].map({"YES":"1" ,"NO":"0"})
    advanced_options.key = advanced_options.key.str.upper().str.strip()

    advanced_options['key'] = advanced_options['key'].replace({
        'AGG PERIOD SEARCH': "agg_period_search",
        "AGG CONSTRAINT CUTOFF":"agg_constraint_cutoff"
    }) 
    advanced_options_json = dict(zip(advanced_options["key"], advanced_options["value"]))

    # Read Base Settings Sheet
    spec_details_df = base_settings_sheet.iloc[0:8, 0:2]
    spec_details_df.columns = ['key', 'value']
    spec_details_df.key = spec_details_df.key.str.upper().str.strip()
    spec_details_df['key'] = spec_details_df['key'].replace({
        'PROJECT': "project",
        "MODEL KEY":"model_key",
        'SPEC TYPE': "spec_type",
        'SPEC VERSION': "spec_version",
        'SPEC ALIAS': "spec_alias",
        'SPEC DESCRIPTION': "spec_description",
        'ADS VERSION': "ads_version",
        'ADS ALIAS': "ads_alias" 
        })

    spec_details_json = dict(zip(spec_details_df["key"], spec_details_df["value"]))

    # Ads Identifiers Table
    ads_identifiers_df = base_settings_sheet.iloc[10:13, 0:2]
    ads_identifiers_df.columns = ['key', 'value']
    ads_identifiers_df["key"] = ads_identifiers_df["key"].str.upper().str.strip()
    # Renaming of keys
    ads_identifiers_df.loc[ads_identifiers_df.key == "DEPENDENT", 'key'] = "dependent_colname" 
    ads_identifiers_df.loc[ads_identifiers_df.key == "GEOGRAPHY", 'key'] = "geo_colname" 
    ads_identifiers_df.loc[ads_identifiers_df.key == "PERIOD", 'value'] =  pd.to_datetime(ads_identifiers_df[ads_identifiers_df["key"] == "PERIOD"]["value"])
    ads_identifiers_df.loc[ads_identifiers_df.key == "PERIOD", 'key'] = "period_colname" 
    ads_identifiers_json = dict(zip(ads_identifiers_df["key"], ads_identifiers_df["value"]))

    # Options Table
    options_df = base_settings_sheet.iloc[15:23, 0:2]
    options_df.columns = ['key', 'value']
    options_df['value'] = options_df["value"].replace(to_replace=["NO", "YES"], value=[str(0), str(1)])

    options_df.key = options_df.key.str.upper().str.strip()

    options_df['key'] = options_df['key'].replace({
        'INCLUDE INTERCEPT': "include_intercept",
        "SEASONALITY CYCLE":"seasonality_cycle",
        'SEASONALITY TYPE': "seasonality_variable",
        'STOCHASTIC SEASONALITY': "include_seasonality",
        'SINGLE TRANSFORMATION': "single_transformation",
        'EXCLUDE DEPENDENT ZERO': "exclude_dependent_zero",
        'FAST OPTIMIZATION': "fast_opt",
        'RAPID REFRESH': "rapid_refresh"
        })

    options_json = dict(zip(options_df["key"], options_df["value"]))


    # Period Definition Table
    period_definition_df = base_settings_sheet.iloc[1:, 3:6]
    period_definition_df.columns = ['name', 'start_date', 'end_date']
    period_definition_df = period_definition_df.dropna()
    period_definition_df['start_date'] = pd.to_datetime(period_definition_df['start_date'])
    period_definition_df['end_date'] = pd.to_datetime(period_definition_df['end_date'])

    transformation = period_definition_df[period_definition_df['name'] == 'TRANSFORMATION']
    modelling = period_definition_df[period_definition_df['name'] == 'MODELLING']
    modelling = modelling.drop('name', 1)
    modelling = modelling.to_dict(orient='records')
    modelling_period_json = {"modelling_period":modelling[0]}

    # period_definition_df['name'] = period_definition_df['name'].replace({'TRANSFORMATION': "transformation","MODELLING":"modelling"})
    period_definition_array = period_definition_df.to_dict(orient='records')
    period_definition_json = {"period_definition":period_definition_array}

    # Reading Reporting Period
    reporting_period_df = base_settings_sheet.iloc[1:, 7:9]
    reporting_period_df.columns = ['shortname', 'name']
    reporting_period_df = reporting_period_df.dropna()

    diff = set(reporting_period_df.name.values).difference(set(pd.merge(reporting_period_df, period_definition_df[['name', 'start_date', 'end_date']],on='name').name.values))
    # reporting_period_df = pd.merge(reporting_period_df, period_definition_df[['name', 'start_date', 'end_date']],on='name')
    if len(diff) >=1:
        raise Exception(f'Reporting Period {diff} are missing')
    else:
        reporting_period_df = pd.merge(reporting_period_df, period_definition_df[['name', 'start_date', 'end_date']],on='name')
        
    reporting_period_df['start_date'] = pd.to_datetime(reporting_period_df['start_date'])
    reporting_period_df['end_date'] = pd.to_datetime(reporting_period_df['end_date'])
    reporting_period_df = reporting_period_df.drop('shortname', 1)

    # reporting_period_df = reporting_period_df.rename(columns = reporting_period_columns)
    reporting_period_df = reporting_period_df.dropna()
    reporting_period_array = reporting_period_df.to_dict(orient='records')
    pval_config_json = {"pval_config":config_json}
    reporting_period_json = {"reporting_period":reporting_period_array}
    transformation_period_json = {"transformation_period":pd.to_datetime(modelling[0]["start_date"])}

    # combining all small json 
    basic_setting_json = {'basic_settings':{**options_json, **ads_identifiers_json, **transformation_period_json, **modelling_period_json, **period_definition_json, **pval_config_json, **reporting_period_json, **seasonality_exclude_json, **advanced_options_json},'configuration':business_constraints_array}

    json_path = xlsx_path.split(".")[0] + ".json"
    with open(json_path, 'w') as fp:
        spec_json = json.dumps(basic_setting_json, ignore_nan=True, default=datetime.datetime.isoformat)
        fp.write(spec_json)

        return(spec_details_json)

# convert json to xlsx
def as_text(value):
    if value is None:
        return ""
    return str(value)

def json_to_xlsx(json_path, output_path, spec_obj):
    json_path = "\\".join([path_to_core, json_path])
    output_path = "\\".join([path_to_core, output_path])
    with open(json_path) as json_file:
        spec = json.load(json_file)
        
    wb = xl.Workbook()
    #--------------------------------- Business Constraints ---------------------------#
    contribution_columns = {
        'include':'Include',
        'type':'Type',
        'variable':'Variable',
        'desp':'Variable Description',
        'contri_period':"Period",
        'min':"Min (%)",
        'max':'Max (%)',
        'sign':'Sign',
        'transformation':'Transformation',
        'transformation_values':'Trans. Values',
        'postmul':'Post Multiplication',
        'dependent':'Dependent'
    }
    ws = wb.create_sheet("Business Constraints", 1)
    business_constraints_df = pd.DataFrame(spec['configuration'])

    business_constraints_df = business_constraints_df.reindex(columns = list(contribution_columns.keys()))
    business_constraints_df = business_constraints_df.rename(columns = contribution_columns)
    business_constraints_df['Include'] = business_constraints_df['Include'].map({1:"YES" , 0:"NO"})

    #--------------------------------- STYLE ---------------------------#
    fill_orange = xl.styles.PatternFill(start_color='f79646', end_color='f79646', fill_type='solid')
    fill_light_orange = xl.styles.PatternFill(start_color='fde9d9', end_color='fde9d9', fill_type='solid')
    fill_light_green  = xl.styles.PatternFill(start_color='ebf1de', end_color='ebf1de', fill_type='solid')
    fill_light_grey = xl.styles.PatternFill(start_color='d9d9d9', end_color='d9d9d9', fill_type='solid')
    ft = xl.styles.Font(color="000000",bold=True)
    thin_border = xl.styles.borders.Border(left=xl.styles.borders.Side(style='thin'), right=xl.styles.borders.Side(style='thin'), top=xl.styles.borders.Side(style='thin'), bottom=xl.styles.borders.Side(style='thin'))
    thin_border_side = xl.styles.borders.Border(left=xl.styles.borders.Side(style='thin'), right=xl.styles.borders.Side(style='thin'))

    for r in dataframe_to_rows(business_constraints_df, index=False, header=True):
        ws.append(r)

    for cell2 in ws[1]:
        cell2.fill = fill_orange
        cell2.font = ft
        cell2.border = thin_border

    for col_cells in ws.iter_cols(min_col=5, max_col=7, min_row=2):
        for cell in col_cells:
            cell.fill = fill_light_orange
            cell.border = thin_border

    for col_cells in ws.iter_cols(min_col=9, max_col=10, min_row=2):
        for cell in col_cells:
            cell.fill = fill_light_green
            cell.border = thin_border

    for col_cells in ws.iter_cols(min_col=3, max_col=3, min_row=2):
        for cell in col_cells:
            cell.font = xl.styles.Font(color="000000",bold=True)

    # Percentage format for min and max
    for col_cells in ws.iter_cols(min_col=6, max_col=7, min_row=1):
        for cell in col_cells:
            cell.number_format = '0.0000%'

    # column width calculation
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length+2

    #--------------------------------- CONFIG ---------------------------#
    configuration_data = spec['basic_settings']['pval_config']
    configuration_data2 = {"Definition":list(configuration_data.keys()), "Value": list(configuration_data.values())}  #coverting list of string to list of int
    configuration_df = pd.DataFrame(configuration_data2)
    configuration_df.index = configuration_df.index + 2 
    configuration_df.loc[0] = ['Pval Config',""]
    configuration_df.loc[1] = ['Definition','Value']
    configuration_df = configuration_df.sort_index()
    # seasonality_exclude_list = spec['basic_settings']['seasonality_exclude']
    seasonality_exclude_list = [1,3,4,56,7,6,5,4] #seasonality_exclude
    seasonality_df = pd.DataFrame(seasonality_exclude_list, columns=['seasonality_exclude'])
    seasonality_df.index = seasonality_df.index + 1 
    seasonality_df.loc[0] = ['Seasonality Exclude Dates (Locations)']
    seasonality_df = seasonality_df.sort_index()


    advance_options_data = {"Agg Period Search":spec['basic_settings']['agg_period_search'],"Agg Constraint Cutoff":spec['basic_settings']['agg_constraint_cutoff'] }
    advance_options_df = pd.DataFrame(advance_options_data.items(), columns=['key', 'value'])
    advance_options_df.index = advance_options_df.index + 1 
    advance_options_df.loc[0] = ['Advance Options',""]
    advance_options_df = advance_options_df.sort_index()
    advance_options_df["value"] = advance_options_df["value"].replace(to_replace=[0, 1], value=["NO", "YES"])

    configuration_df_shape = configuration_df.shape
    seasonality_df_shape = seasonality_df.shape
    advance_options_df_shape = advance_options_df.shape

    parent_config_df=np.full((30,6),np.nan)
    parent_config_df=pd.DataFrame(parent_config_df)

    parent_config_df.iloc[0:configuration_df_shape[0],0:2] =configuration_df.values
    parent_config_df.iloc[0:seasonality_df_shape[0],2:3] = seasonality_df.values
    parent_config_df.iloc[0:advance_options_df_shape[0],4:6] = advance_options_df.values

    ws2 = wb.create_sheet("Config", 2)
    for each_row in dataframe_to_rows(parent_config_df, index=False, header=False):
        ws2.append(each_row)

    for column_cells in ws2.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws2.column_dimensions[column_cells[0].column_letter].width = length+2

    for col_cells in ws2.iter_cols(min_col=1, max_col=2, min_row=1, max_row=1):
        for cell in col_cells:
            cell.fill = fill_orange
            cell.font = ft

    for col_cells in ws2.iter_cols(min_col=3, max_col=3, min_row=1, max_row=1):
        for cell in col_cells:
            cell.fill = fill_orange
            cell.font = ft
            cell.border = thin_border

    for col_cells in ws2.iter_cols(min_col=3, max_col=3, min_row=2):
        for cell in col_cells:
            cell.border = thin_border_side
    
    for col_cells in ws2.iter_cols(min_col=5, max_col=6, min_row=1, max_row=1):
        for cell in col_cells:
            cell.fill = fill_orange
            cell.font = ft

    #--------------------------------- Base Settings ---------------------------#

    reporting_data = spec['basic_settings']['reporting_period']
    reporting_df = pd.DataFrame(reporting_data)

    reporting_df_period = reporting_df[["name", 'start_date', 'end_date' ]]
    reporting_df_period = reporting_df_period.rename(columns = {'name':'Definition','start_date':'Start', 'end_date':'End'})

    reporting_df=reporting_df.rename(columns = {'name':'Shortname'})
    reporting_df['Definition'] = reporting_df['Shortname']

    # Reporting Dataframe
    reporting_df = reporting_df[['Shortname','Definition']]
    reporting_df.reset_index(drop=True, inplace = True)
    reporting_df.index  = reporting_df.index + 2
    reporting_df.loc[0]=['REPORTING PERIOD','']
    reporting_df.loc[1]=['Shortname','Definition']
    reporting_df = reporting_df.sort_index()

    # Period Definition
    transformation_period = spec['basic_settings']['transformation_period']
    modelling_period = spec['basic_settings']['modelling_period']
    periodDefData = [
        ['TRANSFORMATION', transformation_period, modelling_period['end_date']],
        ['MODELLING',  modelling_period['start_date'], modelling_period['end_date']]]

    period_definition_df = pd.DataFrame(periodDefData,columns = ['Definition', 'Start','End'])
    period_definition_df = pd.concat([period_definition_df, reporting_df_period]) # adding TRANSFORMATION and modelling period with reporting period(reporting_df_period)
    period_definition_df['Start'] = pd.to_datetime(period_definition_df['Start']).dt.strftime("%m-%d-%Y")
    period_definition_df['End'] = pd.to_datetime(period_definition_df['End']).dt.strftime("%m-%d-%Y")
    period_definition_df.reset_index(drop=True, inplace = True)
    period_definition_df.index = period_definition_df.index + 2 
    period_definition_df.loc[0] = ['PERIOD DEFINITIONS','','']
    period_definition_df.loc[1] = ['Definition','Start','End']
    period_definition_df = period_definition_df.sort_index()


    # Options Table
    options_data = [
        ['OPTIONS', ""],
        ['Include Intercept', spec['basic_settings']['include_intercept']],
        ['Seasonality Type',  spec['basic_settings']['seasonality_variable']],
        ['Seasonality Cycle', spec['basic_settings']['seasonality_cycle']],
        ['Stochastic Seasonality',  spec['basic_settings']['stochastic_sea']],
        ['Single Transformation', spec['basic_settings']['single_transformation']],
        ['Exclude Dependent Zero',  spec['basic_settings']['exclude_dependent_zero']],
        ['Fast Optimization', spec['basic_settings']['fast_opt']],
        ['Rapid Refresh',  spec['basic_settings']['exclude_dependent_zero']]
    ]
    option_df = pd.DataFrame(options_data, columns=['OPTIONS', ''])
    option_df[""] = option_df[""].replace(to_replace=[0, 1], value=["NO", "YES"])
    # Ads Identifiers Table

    ads_identifiers_data = [
    ['ADS IDENTIFIERS',""],
    ['Dependent',spec['basic_settings']['dependent_colname']], 
    ['Geography',  spec['basic_settings']['geo_colname']],
    ['Period',datetime.datetime.strptime(spec['basic_settings']['transformation_period'],"%Y-%m-%dT%H:%M:%S").strftime("%m-%d-%Y")],
    ]
        
    ads_identifiers_df = pd.DataFrame(ads_identifiers_data,columns = ['ADS IDENTIFIERS', ''])

    spec_details_data = [["SPEC DETAILS",""],
    ["Project",spec_obj["project"]],
    ["Model Key",spec_obj["model_key"]],
    ["Spec Type",spec_obj["spec_type"]],
    ["Spec Version",spec_obj["spec_version"]],
    ["Spec Alias",spec_obj["spec_alias"]],
    ["Spec Description",spec_obj["spec_description"]],
    ["ADS Version",spec_obj["ads_version"]],
    ["ADS Alias",spec_obj["ads_alias"]]
    ]
    spec_details_df = pd.DataFrame(spec_details_data,columns = ['SPEC DETAILS', ''])

    # Dataframes shape size
    l = period_definition_df.shape
    rp = reporting_df.shape
    sd = spec_details_df.shape
    ai = ads_identifiers_df.shape
    op = option_df.shape

    a=np.full((30,10),np.nan)
    a=pd.DataFrame(a)

    a.iloc[0:l[0],3:6] =period_definition_df.values

    a.iloc[0:rp[0],7:9] = reporting_df.values
    a.iloc[0:sd[0],0:2] = spec_details_df.values
    a.iloc[sd[0]+1:sd[0]+ai[0]+1,0:2] = ads_identifiers_df.values
    a.iloc[sd[0]+ai[0]+2:sd[0]+ai[0]+op[0]+2,0:2] = option_df.values

    ws3 = wb.create_sheet("Base Settings", 0)

    #--------------------------------- Styles - Base Settings ---------------------------#

    for each_row in dataframe_to_rows(a, index=False, header=False):
        ws3.append(each_row)

    for col_cells in ws3.iter_cols(min_col=1, max_col=2, min_row=1, max_row=1):
        for cell in col_cells:
            cell.fill = fill_orange
            cell.font = ft

    for col_cells in ws3.iter_cols(min_col=1, max_col=2, min_row=2, max_row=sd[0]):
        for cell in col_cells:
            cell.border = thin_border

    for col_cells in ws3.iter_cols(min_col=1, max_col=2, min_row=11, max_row=11):
        for cell in col_cells:
            cell.fill = fill_orange
            cell.font = ft 

    for col_cells in ws3.iter_cols(min_col=1, max_col=2, min_row=12, max_row=(12+ai[0]-2)):
        for cell in col_cells:
            cell.border = thin_border


    for col_cells in ws3.iter_cols(min_col=1,max_col=2, min_row=16, max_row=16):
        for cell in col_cells:
            cell.fill = fill_orange
            cell.font = ft 

    for col_cells in ws3.iter_cols(min_col=1, max_col=2, min_row=17, max_row=(17+op[0]-2)):
        for cell in col_cells:
            cell.border = thin_border

    for col_cells in ws3.iter_cols(min_col=4, max_col=6, min_row=1, max_row=1):
        for cell in col_cells:
            cell.fill = fill_orange
            cell.font = ft

    for col_cells in ws3.iter_cols(min_col=4, max_col=6, min_row=2, max_row=2):
        for cell in col_cells:
            cell.fill = fill_light_grey
            cell.border = thin_border
            cell.font = ft

    for col_cells in ws3.iter_cols(min_col=4, max_col=6, min_row=3, max_row=4):
        for cell in col_cells:
            cell.fill = fill_light_orange
            cell.border = thin_border

    transformation_cell = ws3.cell(3, 4)
    transformation_cell.font = ft

    modelling_cell = ws3.cell(4, 4)
    modelling_cell.font = ft

    for col_cells in ws3.iter_cols(min_col=4, max_col=6, min_row=3, max_row=l[0]):
        for cell in col_cells:
            cell.border = thin_border

    for col_cells in ws3.iter_cols(min_col=8, max_col=9, min_row=1, max_row=1):
        for cell in col_cells:
            cell.fill = fill_orange
            cell.font = ft 

    for col_cells in ws3.iter_cols(min_col=8, max_col=9, min_row=2, max_row=2):
        for cell in col_cells:
            cell.fill = fill_light_grey
            cell.border = thin_border
            cell.font = ft 

    for col_cells in ws3.iter_cols(min_col=8, max_col=9, min_row=2, max_row=rp[0]):
        for cell in col_cells:
            cell.border = thin_border

    for column_cells in ws3.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws3.column_dimensions[column_cells[0].column_letter].width = length

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
        
    wb.save(output_path)

def convert_bucket_aggregation(path):   
    print("path main, ", path)
    columns = list(pd.read_csv(path, low_memory=False, memory_map=True, engine='c').head(0))
    print("columns",len(columns))
    data = pd.read_csv(path, low_memory=False, memory_map=True, engine='c')[columns]
    data.columns = columns = [k for k, l in enumerate(columns)]# converting headers to 0, 1, 2, 3, ....
    out = []
    for i, j in enumerate(columns):
        print(i,j)
        rand_mapping_dict = data.set_index(columns[i]).T.to_dict('dict')  
        for k,v in rand_mapping_dict.items():
            obj = {}
            obj["Level"] = i+1
            obj["Name"] = k
            obj["Parent_Name"] = rand_mapping_dict[k][0+i+1] if i!=columns[len(columns)-1] else "" 
            out.append(obj)  
            df = pd.DataFrame(out) 
            df.to_csv(path, index=False)
    print("Aggr template done")                